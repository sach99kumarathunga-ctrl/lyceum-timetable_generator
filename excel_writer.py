"""
excel_writer.py  –  Excel read/write for Lyceum Timetable Generator
Supports:
  - "flexible" mode (Grades 6/7/8): new per-class-period Setup sheet
  - "legacy"   mode (Grades 9/10/11-12): original stream-based Setup sheet
"""
import os, re
from copy import copy
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.datavalidation import DataValidation
from scheduler import TIME_SLOTS, PERIOD_SNS, SLOT_ROW, DAYS, ASSEMBLY_DAY, ASSEMBLY_PERIOD

_HERE = os.path.dirname(os.path.abspath(__file__))

thin = Side(style="thin",   color="CCCCCC")
med  = Side(style="medium", color="999999")
STD  = Border(left=thin, right=thin, top=thin, bottom=thin)
MED  = Border(left=med,  right=med,  top=med,  bottom=med)

DAY_COLORS  = ["1F3864","2E75B6","375623","843C0C","4A235A"]
SUBJ_COLORS = {
    "Mathematics":"BDD7EE","English":"E2EFDA","Science":"C6EFCE",
    "Sinhala":"FFF2CC","ICT":"E1D5E7","History":"FCE4D6","PE":"C6EFCE",
    "Religion":"FFF2CC","Art":"FFE6CC","Dancing":"FCE4D6","Music":"FFF2CC",
    "Chemistry":"FFF2CC","Biology":"C6EFCE","Physics":"FCE4D6",
    "Cookery":"FFE6CC","Lifeskill":"E2EFDA","Civic":"DAE3F3",
    "Chinese":"E0F0FF","French":"E8E0F8","History/Geo":"FCE4D6",
    "Effective Speech":"E2EFDA","Speech and Drama":"E2EFDA",
    "Buddhism":"FFF2CC","Buddhism/PE":"FFF2CC","Assembly":"E8E8E8",
}
DEFAULT_COLOR = "F0F0F0"

def _subj_color(name):
    for k, v in SUBJ_COLORS.items():
        if k.lower() in name.lower():
            return v
    return DEFAULT_COLOR

def cs(ws, r, c, val="", bold=False, bg=None, fg="000000", sz=10,
       al="center", va="center", wrap=False, bdr=STD, italic=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz, italic=italic)
    if bg: cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=al, vertical=va, wrap_text=wrap)
    if bdr: cell.border = bdr
    return cell

def mhdr(ws, r, c1, c2, val, bg="1F3864", fg="FFFFFF", sz=11, h=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cs(ws, r, c1, val, bold=True, bg=bg, fg=fg, sz=sz, al="left")
    if h: ws.row_dimensions[r].height = h

# ══════════════════════════════════════════════════════════════
#   FLEXIBLE SETUP SHEET (Grades 6, 7, 8)
# ══════════════════════════════════════════════════════════════
# Column layout (1-indexed):
#  A=1  Teacher Name
#  B=2  Subject
#  C=3  Room
#  D=4  Same Day (YES/NO)
#  E=5  All Classes 1 Period (YES/NO)
#  F=6  Merge Groups  (e.g. "6A+6B; 6C+6D; 6E+6F" )
#  G=7  [blank spacer]
#  H..  class columns (one per class)

FLEX_COL_TEACHER  = 1
FLEX_COL_SUBJECT  = 2
FLEX_COL_ROOM     = 3
FLEX_COL_SAMEDAY  = 4
FLEX_COL_ALL1     = 5
FLEX_COL_MERGE    = 6
FLEX_COL_SPACER   = 7
FLEX_COL_CLASSES  = 8   # classes start at col 8

def _add_yesno_dropdown(ws, col_letter, row_start, row_end):
    """Add Yes/No dropdown validation to a column range."""
    dv = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False,
        showDropDown=False,   # False = show the arrow button in Excel
        showErrorMessage=True,
        errorTitle="Invalid",
        error="Please select Yes or No from the dropdown.",
    )
    dv.sqref = f"{col_letter}{row_start}:{col_letter}{row_end}"
    ws.add_data_validation(dv)


def write_setup_flexible(ws, grp):
    """Write the new flexible Setup sheet for Grades 6/7/8 with Yes/No dropdowns."""
    ws.sheet_view.showGridLines = False
    classes    = grp["classes"]
    nc         = len(classes)
    total_cols = FLEX_COL_CLASSES + nc + 1
    day_name   = ["Monday","Tuesday","Wednesday","Thursday","Friday"][ASSEMBLY_DAY - 1]

    # ── Header rows ──
    mhdr(ws,1,1,total_cols,
         f"  {grp['name'].upper()} TIMETABLE  |  Edit here → Save → Generate",
         "1F3864", sz=13, h=34)
    mhdr(ws,2,1,total_cols,
         f"  ✏  Set periods per class in the class columns. "
         f"Use the dropdowns in Same Day / All Classes columns. "
         f"Merge Groups: e.g. 6A+6B; 6C+6D  |  "
         f"📌 Assembly is always fixed: {day_name} Period 1 (7:40–8:15)",
         "2E75B6", sz=9, h=28)
    mhdr(ws,3,1,total_cols,
         "  ⚠  Same Day=Yes → all classes get this subject on the same weekday.  "
         "All Classes 1 Period=Yes → 1 period per every class, same day.  "
         "Set periods to 0 to skip a class.",
         "C00000", fg="FFFFFF", sz=9, h=16)
    ws.row_dimensions[4].height = 6

    # ── Column header row (row 5) ──
    hdrs = ["Teacher Name","Subject","Room","Same Day?","All Classes\n1 Period?","Merge Groups\n(e.g. 6A+6B; 6C+6D)",""]
    for ci, h in enumerate(hdrs, 1):
        cs(ws,5,ci,h,bold=True,bg="1F3864",fg="FFFFFF",sz=9,wrap=True)
    for i, cls in enumerate(classes):
        cs(ws,5,FLEX_COL_CLASSES+i,cls,bold=True,bg="2E75B6",fg="FFFFFF",sz=9)
    ws.row_dimensions[5].height = 30

    # ── Teacher reference table (right side) ──
    tbl_col = FLEX_COL_CLASSES + nc + 2
    mhdr(ws,5,tbl_col,tbl_col+2,"  TEACHERS (reference)","375623",fg="FFFFFF",sz=9,h=30)
    for ci,h in enumerate(["Name","Subject"],tbl_col):
        cs(ws,6,ci,h,bold=True,bg="C6EFCE",sz=9)
    for ri,(tid,tname,tsubj) in enumerate(grp["teachers"], 7):
        bg = "FFFFFF" if ri%2 else "F5F5F5"
        cs(ws,ri,tbl_col,  tname,bg=bg,al="left",sz=9)
        cs(ws,ri,tbl_col+1,tsubj,bg=bg,al="left",sz=9)
        ws.row_dimensions[ri].height = 15

    # ── Subject rows ──
    data_row_start = 7
    row = data_row_start
    for srow in grp.get("subject_rows", []):
        # Handle religion_block dict entries specially
        if isinstance(srow, dict) and srow.get("type") == "religion_block":
            # Header row for the block
            cs(ws,row,FLEX_COL_TEACHER,"⛪ RELIGION BLOCK — All classes merge (1 period/week)",
               bg="FFE0B2",al="left",sz=9,bold=True)
            cs(ws,row,FLEX_COL_SUBJECT,
               "Buddhism: multiple teachers (high count). Islam/Hindu/RC/Christianity: 1 teacher each (low count). ALL same slot.",
               bg="FFE0B2",al="left",sz=8,italic=True,wrap=True)
            for c in range(FLEX_COL_ROOM, FLEX_COL_CLASSES+len(classes)+1):
                cs(ws,row,c,"",bg="FFE0B2")
            ws.row_dimensions[row].height = 18
            row += 1
            # One row per religion teacher
            for t in srow.get("teachers",[]):
                bg2 = "FFF8F0" if row%2==0 else "FFF3E0"
                covers = t.get("covers", "All classes")
                cs(ws,row,FLEX_COL_TEACHER, f"  ↳ {t['teacher']}",bg=bg2,al="left",sz=9)
                cs(ws,row,FLEX_COL_SUBJECT, t["subject"],           bg=bg2,al="left",sz=9,bold=True)
                cs(ws,row,FLEX_COL_ROOM,    t.get("room","R-REL"),  bg=bg2,al="center",sz=9)
                cs(ws,row,FLEX_COL_SAMEDAY, "Yes",                  bg="FFFDE7",al="center",sz=9)
                cs(ws,row,FLEX_COL_ALL1,    "Yes",                  bg="E8F5E9",al="center",sz=9)
                cs(ws,row,FLEX_COL_MERGE,   covers,                 bg=bg2,al="left",sz=9,wrap=True)
                cs(ws,row,FLEX_COL_SPACER,  "",                     bg=bg2)
                # Mark all class columns with 1 (whole grade attends)
                for i2, cls2 in enumerate(classes):
                    cs(ws,row,FLEX_COL_CLASSES+i2,"1",bg="FFE0B2",sz=9,al="center")
                ws.row_dimensions[row].height = 15
                row += 1
            row += 1   # gap after religion block
            continue

        # Handle merge_block dicts — display each teacher as a row with per-class periods
        if isinstance(srow, dict) and srow.get("type") == "merge_block":
            mb     = srow
            pairs  = mb.get("merge_pairs", [])
            std    = mb.get("standalone", {})
            period = mb.get("periods", 2)
            # Build per-class period counts for merged classes
            merged_cls = [c for pair in pairs for c in pair if c in classes]
            # Colour bands matching Clive sheet: yellow=pair1, blue=pair2, green=pair3
            pair_bgs = ["FFFF00","BDD7EE","E2EFDA","FCE4D6","DAE3F3"]

            # Header row for the block
            pairs_str = "; ".join("+".join(p) for p in pairs)
            cs(ws,row,FLEX_COL_TEACHER,
               f"🔀 MERGE BLOCK — {pairs_str}  (same day, same slot, students choose subject)",
               bg="FFF9C4",al="left",sz=9,bold=True)
            cs(ws,row,FLEX_COL_SUBJECT,"All 3 teachers active simultaneously per pair",
               bg="FFF9C4",al="left",sz=8,italic=True)
            for c in range(FLEX_COL_ROOM, FLEX_COL_CLASSES+len(classes)+1):
                cs(ws,row,c,"",bg="FFF9C4")
            ws.row_dimensions[row].height=15; row+=1

            # One row per teacher in the block
            for t in mb.get("teachers",[]):
                bg_t = "FFFDE7" if row%2==0 else "FFF9C4"
                # Per-class period dict: merged classes get `period`, standalone get their value, others 0
                pc_display = {}
                for cls2 in classes:
                    if any(cls2 in pair for pair in pairs):
                        pc_display[cls2] = period
                    elif cls2 in std:
                        pc_display[cls2] = std[cls2]
                    else:
                        pc_display[cls2] = 0
                merge_str_t = "; ".join("+".join(p) for p in pairs)

                cs(ws,row,FLEX_COL_TEACHER, "🔀 "+t["teacher"], bg=bg_t, al="left",   sz=9)
                cs(ws,row,FLEX_COL_SUBJECT, t["subject"],   bg=bg_t, al="left",   sz=9, bold=True)
                cs(ws,row,FLEX_COL_ROOM,    t.get("room","R-MRG"), bg=bg_t, al="center", sz=9)
                cs(ws,row,FLEX_COL_SAMEDAY, "Yes",          bg="FFFDE7", al="center", sz=9)
                cs(ws,row,FLEX_COL_ALL1,    "No",           bg=bg_t,  al="center", sz=9)
                cs(ws,row,FLEX_COL_MERGE,   merge_str_t,    bg="E3F2FD", al="left",  sz=9, wrap=True)
                cs(ws,row,FLEX_COL_SPACER,  "",             bg=bg_t)

                for i2, cls2 in enumerate(classes):
                    pw = pc_display.get(cls2, 0)
                    # Colour-code by merge pair
                    pair_idx = next((pi for pi,pair in enumerate(pairs) if cls2 in pair), None)
                    if pair_idx is not None:
                        cbg = pair_bgs[pair_idx % len(pair_bgs)]
                    elif cls2 in std:
                        cbg = "E8E8E8"
                    else:
                        cbg = bg_t
                    cs(ws,row,FLEX_COL_CLASSES+i2, pw if pw else "", bg=cbg, sz=9, al="center",
                       bold=(pair_idx is not None))
                ws.row_dimensions[row].height=15; row+=1
            row+=1   # gap after block
            continue

        # Normal tuple row
        if isinstance(srow, dict):
            continue  # skip any other unhandled dict types
        tch  = srow[0]; subj = srow[1]
        room = srow[6] if len(srow)>6 else "R??"
        pc   = srow[2]; mgs = srow[3]
        sd   = srow[4]; a1  = srow[5]
        merge_str = "; ".join("+".join(mg) for mg in mgs) if mgs else ""
        bg = "FFFFFF" if row%2==0 else "F9F9F9"

        cs(ws,row,FLEX_COL_TEACHER, tch,      bg=bg,al="left",  sz=9)
        cs(ws,row,FLEX_COL_SUBJECT, subj,     bg=bg,al="left",  sz=9)
        cs(ws,row,FLEX_COL_ROOM,    room,     bg=bg,al="center",sz=9)
        # Dropdown cells — value set, dropdown applied later
        sd_val = "Yes" if sd else "No"
        a1_val = "Yes" if a1 else "No"
        cell_sd = ws.cell(row=row, column=FLEX_COL_SAMEDAY, value=sd_val)
        cell_sd.font      = Font(name="Arial", size=9,
                                  bold=True if sd else False,
                                  color="1A5276" if sd else "555555")
        cell_sd.fill      = PatternFill("solid", start_color="FFFDE7" if sd else bg)
        cell_sd.alignment = Alignment(horizontal="center", vertical="center")
        cell_sd.border    = STD

        cell_a1 = ws.cell(row=row, column=FLEX_COL_ALL1, value=a1_val)
        cell_a1.font      = Font(name="Arial", size=9,
                                  bold=True if a1 else False,
                                  color="1A5276" if a1 else "555555")
        cell_a1.fill      = PatternFill("solid", start_color="E8F5E9" if a1 else bg)
        cell_a1.alignment = Alignment(horizontal="center", vertical="center")
        cell_a1.border    = STD

        cs(ws,row,FLEX_COL_MERGE,  merge_str,bg="E3F2FD" if merge_str else bg,al="left",sz=9,wrap=True)
        cs(ws,row,FLEX_COL_SPACER, "",       bg=bg)
        for i,cls in enumerate(classes):
            pw  = pc.get(cls,0) if pc else 0
            cbg = _subj_color(subj) if pw>0 else bg
            cs(ws,row,FLEX_COL_CLASSES+i, pw if pw else "",bg=cbg,sz=9)
        ws.row_dimensions[row].height = 16
        row += 1

    # ── Blank rows for new entries ──
    blank_start = row
    for _ in range(10):
        bg = "FFFFFF" if row%2==0 else "FFFAED"
        for c in range(1, FLEX_COL_CLASSES+nc+1):
            cs(ws,row,c,"",bg=bg)
        ws.cell(row=row,column=FLEX_COL_SAMEDAY,value="No").font = Font(name="Arial",size=9,color="555555")
        ws.cell(row=row,column=FLEX_COL_SAMEDAY).fill      = PatternFill("solid",start_color=bg)
        ws.cell(row=row,column=FLEX_COL_SAMEDAY).alignment = Alignment(horizontal="center",vertical="center")
        ws.cell(row=row,column=FLEX_COL_SAMEDAY).border    = STD
        ws.cell(row=row,column=FLEX_COL_ALL1,   value="No").font = Font(name="Arial",size=9,color="555555")
        ws.cell(row=row,column=FLEX_COL_ALL1).fill      = PatternFill("solid",start_color=bg)
        ws.cell(row=row,column=FLEX_COL_ALL1).alignment = Alignment(horizontal="center",vertical="center")
        ws.cell(row=row,column=FLEX_COL_ALL1).border    = STD
        ws.row_dimensions[row].height = 16
        row += 1

    data_row_end = row - 1

    # ── Apply Yes/No dropdowns to Same Day and All Classes columns ──
    sd_col_letter = get_column_letter(FLEX_COL_SAMEDAY)
    a1_col_letter = get_column_letter(FLEX_COL_ALL1)
    _add_yesno_dropdown(ws, sd_col_letter, data_row_start, data_row_end)
    _add_yesno_dropdown(ws, a1_col_letter, data_row_start, data_row_end)

    row += 1  # spacer

    # ── Time-slot reference ──
    mhdr(ws,row,1,5,"  TIME SLOTS (reference)","C6EFCE",fg="1F3864",sz=10,h=18)
    row += 1
    for ci,h in enumerate(["Slot","Time","Label","Break?"],1):
        cs(ws,row,ci,h,bold=True,bg="C6EFCE",fg="1F3864",sz=9)
    row += 1
    for sn,tr,lbl,isb in TIME_SLOTS:
        tbg = "FFF2CC" if isb else ("F5F5F5" if row%2 else "FFFFFF")
        asm_note = f"  ← ASSEMBLY FIXED HERE (Every {day_name})" if (not isb and sn==ASSEMBLY_PERIOD) else ""
        cs(ws,row,1,sn,bg=tbg); cs(ws,row,2,tr,bg=tbg,al="left")
        cs(ws,row,3,lbl+asm_note,bg="FFF9C4" if asm_note else tbg,al="left")
        cs(ws,row,4,"YES" if isb else "",bg=tbg)
        ws.row_dimensions[row].height=15
        row += 1

    # ── Column widths ──
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 11   # Same Day
    ws.column_dimensions["E"].width = 13   # All 1 Period
    ws.column_dimensions["F"].width = 22   # Merge Groups
    ws.column_dimensions["G"].width = 3
    for i in range(nc):
        ws.column_dimensions[get_column_letter(FLEX_COL_CLASSES+i)].width = 6


def _read_setup_flexible(ws, grp):
    """
    Read the flexible Setup sheet back into subject_rows format.
    Returns list of dicts ready for schedule_flexible().
    """
    classes = grp["classes"]
    # Find data start (row after the header row 5)
    data_start = 7
    rows = []
    # Pass through dict-type rows from grp data (religion_block, merge_block)
    # These can't be stored in the Setup sheet so always come from grp definition
    for srow in grp.get("subject_rows", []):
        if isinstance(srow, dict) and srow.get("type") in ("religion_block", "merge_block"):
            rows.append(srow)

    for r in range(data_start, data_start + 200):
        tch  = str(ws.cell(r, FLEX_COL_TEACHER).value or "").strip()
        subj = str(ws.cell(r, FLEX_COL_SUBJECT).value or "").strip()
        if not tch or not subj:
            continue
        # Skip religion block display rows written by write_setup_flexible
        if tch.startswith("↳") or tch.startswith("⛪") or "RELIGION BLOCK" in tch or tch.startswith("🔀"):
            continue
        room = str(ws.cell(r, FLEX_COL_ROOM).value or "R??").strip()
        sd_v = str(ws.cell(r, FLEX_COL_SAMEDAY).value or "").strip().upper()
        a1_v = str(ws.cell(r, FLEX_COL_ALL1).value   or "").strip().upper()
        mg_v = str(ws.cell(r, FLEX_COL_MERGE).value  or "").strip()
        sd   = sd_v == "YES"
        a1   = a1_v == "YES"
        # Parse merge groups: "6A+6B; 6C+6D"
        mgs  = []
        if mg_v:
            for part in mg_v.split(";"):
                part = part.strip()
                if "+" in part:
                    mgs.append([c.strip() for c in part.split("+")])
        # Per-class periods
        pc = {}
        for i, cls in enumerate(classes):
            col = FLEX_COL_CLASSES + i
            raw = ws.cell(r, col).value
            try:    pw = int(raw)
            except: pw = 0
            if pw > 0:
                pc[cls] = pw
            # For all_1period rows, fill missing classes with 1
        if a1:
            for cls in classes:
                if cls not in pc:
                    pc[cls] = 1

        if not pc and not a1:
            continue  # skip empty rows

        rows.append({
            "teacher":     tch,
            "subject":     subj,
            "room":        room,
            "per_class":   pc,
            "merge_groups":mgs,
            "same_day":    sd,
            "all_1period": a1,
        })
    return rows


# ══════════════════════════════════════════════════════════════
#   LEGACY SETUP SHEET (Grades 9, 10, 11-12)  — unchanged
# ══════════════════════════════════════════════════════════════
def write_setup_legacy(ws, grp):
    ws.sheet_view.showGridLines = False
    streams  = grp["streams"]
    teachers = grp["teachers"]
    subjects = grp["subjects"]

    mhdr(ws,1,1,15,
         f"  {grp['name'].upper()} TIMETABLE  |  Edit here → Save → Click Generate",
         "1F3864", sz=13, h=34)
    mhdr(ws,2,1,15,
         "  ✏  Change teacher names / periods / rooms below. Do NOT change ID codes.",
         "2E75B6", sz=10, h=20)
    mhdr(ws,3,1,15,
         "  ⚠  To remove a subject set Periods/Wk = 0. To add, fill a blank row.",
         "C00000", fg="FFFFFF", sz=9, h=16)
    ws.row_dimensions[4].height = 8

    if len(streams)==1:
        stream=streams[0]
        mhdr(ws,5,1,6,f"  TEACHERS",grp["stream_colors"][stream],sz=11,h=22)
        hbg="C6EFCE"
        for ci,h in enumerate(["ID","Teacher Name","Subject","Max/Day","Days"],1):
            cs(ws,6,ci,h,bold=True,bg=hbg,sz=9)
        ws.row_dimensions[6].height=18
        for ri,(tid,tname,tsubj,_) in enumerate([t for t in teachers if t[3]==stream],7):
            bg="FFFFFF" if ri%2 else "F5F5F5"
            for ci,v in enumerate([tid,tname,tsubj,5,"Mon-Fri"],1):
                cs(ws,ri,ci,v,bg=bg,al="left" if ci>1 else "center")
            ws.row_dimensions[ri].height=16
        te=7+len([t for t in teachers if t[3]==stream])
    else:
        for si,stream in enumerate(streams):
            sc=1 if stream=="PED" else 8
            label="EDEXCEL / PED STREAM TEACHERS" if stream=="PED" else "NATIONAL OL STREAM TEACHERS"
            mhdr(ws,5,sc,sc+5,f"  {label}",grp["stream_colors"][stream],sz=11,h=22)
            hbg="BDD7EE" if stream=="PED" else "C6EFCE"
            for ci,h in enumerate(["ID","Teacher Name","Subject","Max/Day","Days"],sc):
                cs(ws,6,ci,h,bold=True,bg=hbg,sz=9)
            st=[t for t in teachers if t[3]==stream]
            for ri,(tid,tname,tsubj,_) in enumerate(st,7):
                bg="FFFFFF" if ri%2 else "F5F5F5"
                for ci,v in enumerate([tid,tname,tsubj,5,"Mon-Fri"],sc):
                    cs(ws,ri,ci,v,bg=bg,al="left" if ci>sc else "center")
                ws.row_dimensions[ri].height=16
        te=7+max(len([t for t in teachers if t[3]==s]) for s in streams)

    ws.row_dimensions[te+1].height=8
    ss=te+2

    if len(streams)==1:
        stream=streams[0]
        mhdr(ws,ss,1,7,f"  SUBJECTS (periods/week from PDF)",grp["stream_colors"][stream],sz=11,h=22)
        hbg="FCE4D6"
        for ci,h in enumerate(["Code","Subject Name","Teacher ID","Periods/Wk","Room","Notes"],1):
            cs(ws,ss+1,ci,h,bold=True,bg=hbg,sz=9)
        ws.row_dimensions[ss+1].height=18
        for ri,(code,name,tid,pw,room,notes) in enumerate(subjects[stream],ss+2):
            bg="F5F5F5" if ri%2 else "FFFFFF"
            for ci,v in enumerate([code,name,tid,pw,room,notes],1):
                cs(ws,ri,ci,v,bg=bg,al="left" if ci>1 else "center")
            ws.row_dimensions[ri].height=16
        for ri in range(ss+2+len(subjects[stream]),ss+2+len(subjects[stream])+5):
            for ci in range(1,7): cs(ws,ri,ci,"",bg="FFFAED")
            cs(ws,ri,4,0,bg="FFFAED")
            ws.row_dimensions[ri].height=15
        subj_end=ss+2+len(subjects[stream])+5
    else:
        for si,stream in enumerate(streams):
            sc=1 if stream=="PED" else 9
            label="EDEXCEL/PED SUBJECTS" if stream=="PED" else "NATIONAL OL SUBJECTS"
            mhdr(ws,ss,sc,sc+6,f"  {label}",grp["stream_colors"][stream],sz=11,h=22)
            hbg="FCE4D6" if stream=="PED" else "E8D5F0"
            for ci,h in enumerate(["Code","Subject","Teacher ID","Periods/Wk","Room","Notes"],sc):
                cs(ws,ss+1,ci,h,bold=True,bg=hbg,sz=9)
            for ri,(code,name,tid,pw,room,notes) in enumerate(subjects[stream],ss+2):
                bg="F5F5F5" if ri%2 else "FFFFFF"
                for ci,v in enumerate([code,name,tid,pw,room,notes],sc):
                    cs(ws,ri,ci,v,bg=bg,al="left" if ci>sc else "center")
                ws.row_dimensions[ri].height=16
            for ri in range(ss+2+len(subjects[stream]),ss+2+len(subjects[stream])+4):
                for ci in range(sc,sc+6): cs(ws,ri,ci,"",bg="FFFAED")
                cs(ws,ri,sc+3,0,bg="FFFAED")
                ws.row_dimensions[ri].height=15
        subj_end=ss+2+max(len(v) for v in subjects.values())+6

    ts=subj_end+2
    mhdr(ws,ts,1,4,"  TIME SLOTS (reference)","C6EFCE",fg="1F3864",sz=10,h=18)
    for ci,h in enumerate(["Slot","Time","Label","Break?"],1):
        cs(ws,ts+1,ci,h,bold=True,bg="C6EFCE",fg="1F3864",sz=9)
    for ri,(sn,tr,lbl,isb) in enumerate(TIME_SLOTS,ts+2):
        bg="FFF2CC" if isb else ("F5F5F5" if ri%2 else "FFFFFF")
        cs(ws,ri,1,sn,bg=bg); cs(ws,ri,2,tr,bg=bg,al="left")
        cs(ws,ri,3,lbl,bg=bg,al="left"); cs(ws,ri,4,"YES" if isb else "",bg=bg)
        ws.row_dimensions[ri].height=16

    mhdr(ws,ts,6,9,"  CLASSES & STREAMS","1F3864",sz=10,h=18)
    for ci,h in enumerate(["Class","Stream","Room"],6):
        cs(ws,ts+1,ci,h,bold=True,bg="BDD7EE",sz=9)
    for ri,(cname,cstream) in enumerate(grp["classes"].items(),ts+2):
        bg="F5F5F5" if ri%2 else "FFFFFF"
        cs(ws,ri,6,cname,bg=bg,al="left",bold=True)
        cs(ws,ri,7,cstream,bg=bg)
        cs(ws,ri,8,f"CR-{cname[:4].replace(' ','')}",bg=bg)
        ws.row_dimensions[ri].height=16

    for col,w in zip("ABCDEFGHIJKLMNO",[7,26,18,14,12,14,7,26,18,14,12,14,7,7,7]):
        ws.column_dimensions[col].width=w


def write_setup(ws, grp):
    """Dispatch to flexible or legacy Setup writer."""
    if grp.get("setup_mode") == "flexible":
        write_setup_flexible(ws, grp)
    else:
        write_setup_legacy(ws, grp)


# ══════════════════════════════════════════════════════════════
#   CLASS TIMETABLE SHEET  (flexible & legacy share same output)
# ══════════════════════════════════════════════════════════════
def write_class_sheet_flexible(ws, cls_name, hc, slots_for_class, now_str, n_conf,
):
    """Write one class timetable sheet (flexible mode).
    religion_slot: {(day, period): [(subj, tch, room), ...]} — all religion at that slot.
    """
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    mhdr(ws,1,1,6,f"  {cls_name.upper()}  –  WEEKLY TIMETABLE",hc,sz=13,h=34)
    mhdr(ws,2,1,6,f"  Generated: {now_str}  |  Unplaced: {n_conf}",
         "2E75B6",sz=9,h=18)
    mhdr(ws,3,1,6,
         "  ✏  To change: edit Setup sheet → save → click Generate",
         "595959",fg="FFFFFF",sz=8,h=14)

    # Day headers
    for ci,day in enumerate(DAYS,2):
        cs(ws,4,ci,day,bold=True,bg=DAY_COLORS[ci-2],fg="FFFFFF",sz=10)
    cs(ws,4,1,"Time/Period",bold=True,bg="1F3864",fg="FFFFFF",sz=9)
    ws.row_dimensions[4].height=22

    for sn,tr,lbl,isb in TIME_SLOTS:
        r = sn + 4
        lbl_text = f"{tr}\n{lbl}" if not isb else lbl
        cs(ws,r,1,lbl_text,bg="EEEEEE" if isb else "F0F4FF",sz=8,wrap=True,bold=isb)
        ws.row_dimensions[r].height = 30 if isb else 38
        if isb:
            for ci in range(2,7):
                cs(ws,r,ci,"",bg="F5F5F5")
            continue
        for ci, d in enumerate(range(1,6),2):
            entry = slots_for_class[d].get(sn)
            if entry:
                subj, tch, room = entry
                if tch == "RELIGION_BLOCK":
                    # All classes merged for religion — show cleanly
                    cs(ws,r,ci,"Religion\nPeriod",bg="FFF2CC",sz=9,wrap=True,bold=True)
                elif tch == "ASSEMBLY":
                    cs(ws,r,ci,"Assembly",bg="DCE6F1",sz=9,wrap=True,bold=True)
                elif "MERGE_" in str(tch):
                    pair_lbl = subj.replace("Merge:","") if "Merge:" in str(subj) else "Merge"
                    cs(ws,r,ci,f"{pair_lbl}\nPeriod",bg="FFF9C4",sz=8,wrap=True,bold=True)
                else:
                    txt = f"{subj}\n{tch}\n{room}"
                    cbg = _subj_color(subj)
                    cs(ws,r,ci,txt,bg=cbg,sz=8,wrap=True)
            else:
                cs(ws,r,ci,"",bg="FAFAFA")

    ws.column_dimensions["A"].width = 14
    for col in "BCDEF":
        ws.column_dimensions[col].width = 22


# Legacy class sheet (reused from original)
def write_class_sheet(ws, cls_name, stream, hc, all_slots, all_teachers,
                      class_subjects, class_short, now_str, conflicts):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"
    sched   = all_slots[cls_name]
    n_conf  = sum(1 for (c,_),v in conflicts.items() if c==cls_name and v>0)

    mhdr(ws,1,1,6,f"  {cls_name.upper()}  –  WEEKLY TIMETABLE",hc,sz=13,h=34)
    mhdr(ws,2,1,6,f"  Stream: {stream}  |  Generated: {now_str}  |  Conflicts: {n_conf}",
         "2E75B6",sz=9,h=18)
    mhdr(ws,3,1,6,
         "  ✏  To change: edit Setup sheet → save file → click Generate in the App",
         "595959",fg="FFFFFF",sz=8,h=14)

    for ci,day in enumerate(DAYS,2):
        cs(ws,4,ci,day,bold=True,bg=DAY_COLORS[ci-2],fg="FFFFFF",sz=10)
    cs(ws,4,1,"Time/Period",bold=True,bg="1F3864",fg="FFFFFF",sz=9)
    ws.row_dimensions[4].height=22

    SUBJ_COLORS_LEG = {
        "ENG":"E2EFDA","MAT":"BDD7EE","SCI":"C6EFCE","SIN":"FFF2CC",
        "HIS":"FCE4D6","GEO":"F2DCDB","SOC":"DAE3F3","COM":"D5E8D4",
        "ICT":"E1D5E7","ART":"FFE6CC","PE":"C6EFCE","REL":"FFF2CC",
        "BIO":"C6EFCE","CHE":"FFF2CC","PHY":"FCE4D6","CS":"DAE3F3",
        "CBM":"BDD7EE","ACC":"E2EFDA","BST":"F2DCDB","ECO":"D5E8D4",
        "SPE":"E2EFDA","LIB":"E0E0E0","SDR":"F0E0FF","DAN":"FCE4D6",
        "MUS":"FFF2CC","ASM":"E8E8E8","MOT":"DEEAF1",
    }

    for sn,tr,lbl,isb in TIME_SLOTS:
        r=sn+4
        isbreak = isb
        lbl_full=f"{tr}\n{lbl}"
        cs(ws,r,1,lbl_full if not isbreak else lbl,
           bg="EEEEEE" if isbreak else "F0F4FF",sz=8,wrap=True,bold=isbreak)
        ws.row_dimensions[r].height=30 if isbreak else 34
        if isbreak:
            for ci in range(2,7): cs(ws,r,ci,"",bg="F5F5F5")
            continue
        for ci,d in enumerate(range(1,6),2):
            entry=sched[d].get(sn)
            if entry:
                code,name,tid,room=entry
                tname=all_teachers.get(tid,("?","?",""))[0]
                txt=f"{name}\n{tname}\n{room}"
                cbg=SUBJ_COLORS_LEG.get(code,"F0F0F0")
                cs(ws,r,ci,txt,bg=cbg,sz=8,wrap=True)
            else:
                cs(ws,r,ci,"",bg="FAFAFA")

    ws.column_dimensions["A"].width=14
    for col in "BCDEF": ws.column_dimensions[col].width=20


def write_teacher_overview_flexible(ws, grp, all_slots, now_str):
    ws.sheet_view.showGridLines = False
    classes = grp["classes"]
    hc      = grp.get("stream_color","1F3864")

    mhdr(ws,1,1,7,f"  {grp['name'].upper()}  –  TEACHER OVERVIEW",hc,sz=13,h=34)
    mhdr(ws,2,1,7,f"  Generated: {now_str}","2E75B6",sz=9,h=18)

    r=4
    for cls in classes:
        mhdr(ws,r,1,6,f"  {cls}","2E75B6",sz=11,h=22)
        r+=1
        cs(ws,r,1,"Period",bold=True,bg="1F3864",fg="FFFFFF",sz=9)
        for ci,day in enumerate(DAYS,2):
            cs(ws,r,ci,day,bold=True,bg=DAY_COLORS[ci-2],fg="FFFFFF",sz=9)
        ws.row_dimensions[r].height=20; r+=1
        for sn,tr,lbl,isb in TIME_SLOTS:
            bg="FFF2CC" if isb else ("F5F5F5" if r%2 else "FFFFFF")
            cs(ws,r,1,f"{tr} {lbl}",bg="EEEEEE" if isb else "F0F4FF",sz=8)
            for ci,d in enumerate(range(1,6),2):
                entry=all_slots[cls][d].get(sn)
                if isb: cs(ws,r,ci,"",bg="F5F5F5")
                elif entry:
                    subj,tch,room=entry
                    cs(ws,r,ci,f"{subj}\n{tch}",bg=_subj_color(subj),sz=7,wrap=True)
                else:
                    cs(ws,r,ci,"",bg="FAFAFA")
            ws.row_dimensions[r].height=28 if not isb else 16; r+=1
        r+=1

    ws.column_dimensions["A"].width=16
    for col in "BCDEF": ws.column_dimensions[col].width=22


def write_conflict_log(ws, conflicts, double_booked, now_str):
    ws.sheet_view.showGridLines=False
    mhdr(ws,1,1,6,"  CONFLICT & DOUBLE-BOOKING LOG","C00000",fg="FFFFFF",sz=12,h=28)
    cs(ws,2,1,f"Generated: {now_str}",bg="FCE4D6",sz=9,al="left",bdr=None)
    r=4

    # ── Capacity explanation ─────────────────────────────────
    mhdr(ws,r,1,6,"  ℹ  HOW TO READ THIS LOG","2E75B6",fg="FFFFFF",sz=10,h=20); r+=1
    expl = ("Each class has 40 teaching slots per week (8 periods/day × 5 days). "
            "If total requested periods exceed 40 per class, the excess goes here. "
            "The personal timetables show ONLY what was actually scheduled. "
            "To fix: reduce period counts in the Setup sheet so each class totals ≤ 40.")
    cs(ws,r,1,expl,bg="DEEAF6",sz=9,al="left",wrap=True,bdr=None)
    ws.row_dimensions[r].height=45; r+=2

    if not conflicts and not double_booked:
        cs(ws,r,1,"✅  Zero conflicts and zero double-bookings.",
           bg="C6EFCE",fg="375623",bold=True,sz=11,al="left",bdr=None); return
    if conflicts:
        mhdr(ws,r,1,6,"  UNPLACED PERIODS  (reduce these in Setup sheet)","843C0C",fg="FFFFFF",sz=10,h=20); r+=1
        for ci,h in enumerate(["Class","Subject / Teacher","Periods Unplaced","Action"],1):
            cs(ws,r,ci,h,bold=True,bg="FCE4D6",sz=9); r+=1
        for (cls,code),v in conflicts.items():
            tch_part = code.split("|")[0] if "|" in code else ""
            subj_part = code.split("|")[1] if "|" in code else code
            cs(ws,r,1,cls,bg="FFF2CC",al="left",bold=True)
            cs(ws,r,2,f"{subj_part}  ({tch_part})",bg="FFF2CC",al="left",sz=9)
            cs(ws,r,3,v,bg="FFCCCC",al="center",bold=True)
            cs(ws,r,4,f"Reduce by {v} in Setup sheet",bg="FFF2CC",al="left",sz=8,italic=True)
            r+=1
    if double_booked:
        r+=1
        mhdr(ws,r,1,6,"  TEACHER DOUBLE-BOOKINGS","C00000",fg="FFFFFF",sz=10,h=20); r+=1
        for ci,h in enumerate(["Teacher","Day","Period","Count"],1):
            cs(ws,r,ci,h,bold=True,bg="FCE4D6",sz=9)
        r+=1
        for (tch,d,p),cnt in double_booked.items():
            cs(ws,r,1,str(tch),bg="FFF2CC",al="left")
            cs(ws,r,2,DAYS[d-1],bg="FFF2CC"); cs(ws,r,3,p,bg="FFF2CC"); cs(ws,r,4,cnt,bg="FFF2CC"); r+=1
    ws.column_dimensions["A"].width=10
    ws.column_dimensions["B"].width=34
    ws.column_dimensions["C"].width=16
    ws.column_dimensions["D"].width=28


# ══════════════════════════════════════════════════════════════
#   LEGACY WRITER HELPERS  (teacher_overview, teacher_sheet, etc.)
# ══════════════════════════════════════════════════════════════
def write_teacher_overview(ws, grp, all_slots, all_teachers, stream_teachers, CLASSES):
    ws.sheet_view.showGridLines=False
    streams=grp["streams"]
    hc=grp["stream_colors"].get(streams[0],"1F3864")
    mhdr(ws,1,1,7,f"  {grp['name'].upper()}  –  TEACHER OVERVIEW",hc,sz=13,h=34)
    now_str=datetime.now().strftime("%d %b %Y  %H:%M")
    mhdr(ws,2,1,7,f"  Generated: {now_str}  |  Classes: {len(CLASSES)}",
         "2E75B6",sz=9,h=18)

    SUBJ_C={"ENG":"E2EFDA","MAT":"BDD7EE","SCI":"C6EFCE","SIN":"FFF2CC",
             "HIS":"FCE4D6","COM":"D5E8D4","ICT":"E1D5E7","ART":"FFE6CC",
             "PE":"C6EFCE","REL":"FFF2CC","BIO":"C6EFCE","CHE":"FFF2CC",
             "PHY":"FCE4D6","CBM":"BDD7EE","DAN":"FCE4D6","ASM":"E8E8E8"}
    r=4
    for stream in streams:
        sc=grp["stream_colors"].get(stream,"1F3864")
        label="PED Stream" if stream=="PED" else "National Stream"
        mhdr(ws,r,1,7,f"  {label}",sc,sz=11,h=22); r+=1
        my_cls=[c for c,s in CLASSES.items() if s==stream]
        n=len(my_cls)
        cs(ws,r,1,"Teacher",bold=True,bg="1F3864",fg="FFFFFF",sz=9)
        for ci,cls in enumerate(my_cls,2):
            cs(ws,r,ci,cls,bold=True,bg=DAY_COLORS[(ci-2)%5],fg="FFFFFF",sz=9)
        cs(ws,r,n+2,"Total",bold=True,bg="1F3864",fg="FFFFFF",sz=9)
        ws.row_dimensions[r].height=20; r+=1
        for tid,tname,tsubj,_ in stream_teachers[stream]:
            uname=all_teachers.get(tid,(tname,tsubj,""))[0]
            cs(ws,r,1,uname,bg="F0F4FF",al="left",sz=9)
            tot=0
            for ci,cls in enumerate(my_cls,2):
                cnt=sum(1 for d in range(1,6) for p in PERIOD_SNS
                        if all_slots[cls][d][p] and all_slots[cls][d][p][2]==tid)
                cbg=SUBJ_C.get(tsubj[:3],"F0F0F0")
                cs(ws,r,ci,cnt if cnt else "",bg=cbg if cnt else "FAFAFA",sz=9)
                tot+=cnt
            cs(ws,r,n+2,tot,bg="BDD7EE",bold=True,sz=9)
            ws.row_dimensions[r].height=16; r+=1
        r+=1
    ws.column_dimensions["A"].width=28
    for i in range(1,15):
        ws.column_dimensions[get_column_letter(i+1)].width=12


def write_teacher_sheet(ws, tid, tname, tsubj, stream, school_level,
                        t_sched, class_short, now_str, hc):
    """
    Personal timetable in the EXACT official Lyceum format.
    Colours, fonts, borders, merges, row heights all match Personal_TT_template.xlsx.
    t_sched: {(day_0indexed, period_sn_index): class_name}
    """
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation   = "landscape"

    # ── Exact colours (from template + tint calculation) ─────
    TITLE_BG  = "1E4E79"   # theme:4 tint:-0.5   dark navy blue  (row 1)
    LABEL_BG  = "D9E2F3"   # theme:8 tint:+0.8   light blue      (labels)
    EMP_BG    = "B4C6E7"   # theme:8 tint:+0.6   medium blue     (Employee #)
    NOTES_BG  = "DEEAF6"   # theme:4 tint:+0.8   very light blue (notes)
    BREAK_BG  = "C0C0C0"   # idx:22              silver/grey     (breaks)
    BLACK     = "000000"
    WHITE     = "FFFFFF"

    # ── Border helpers ────────────────────────────────────────
    _thin     = Side(style="thin", color=BLACK)
    _none     = Side(style=None)
    ALL       = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    TB        = Border(top=_thin, bottom=_thin)
    RTB       = Border(right=_thin, top=_thin, bottom=_thin)

    def _f(bold=False, sz=12, color=BLACK, name="Calibri", italic=False):
        return Font(name=name, bold=bold, size=sz, color=color, italic=italic)

    def _fill(hex_color):
        return PatternFill("solid", start_color=hex_color)

    def _aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def put(r, c, val="", font=None, fill=None, aln=None, border=ALL):
        ce = ws.cell(row=r, column=c, value=val)
        if font:   ce.font      = font
        if fill:   ce.fill      = fill
        if aln:    ce.alignment = aln
        if border: ce.border    = border
        return ce

    def merge_put(r, c1, c2, val="", font=None, fill=None, aln=None, h=None):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        ce = ws.cell(row=r, column=c1, value=val)
        if font: ce.font      = font
        if fill: ce.fill      = fill
        if aln:  ce.alignment = aln
        if h:    ws.row_dimensions[r].height = h
        return ce

    # ════════════════════════════════════════════════════════
    # ROW 1 — Title banner  A1:G1  dark navy, white text, centered
    # ════════════════════════════════════════════════════════
    ws.row_dimensions[1].height = 55.95
    merge_put(1, 1, 7,
        "SUBJECT TEACHER TIMETABLE\n"
        "(Prepared as per the official Lyceum International School format)",
        font=_f(bold=False, sz=12, color=WHITE, name="Calibri"),
        fill=_fill(TITLE_BG),
        aln=_aln("center","center",wrap=True))

    # ════════════════════════════════════════════════════════
    # ROW 2 — School name  B2:E2  no fill, bold black, centered
    # ════════════════════════════════════════════════════════
    ws.row_dimensions[2].height = 55.95
    merge_put(2, 2, 5,
        "LYCEUM INTERNATIONAL SCHOOL - Kurunegala",
        font=_f(bold=True, sz=13.5, color=BLACK, name="Calibri"),
        aln=_aln("center","center"))

    # ════════════════════════════════════════════════════════
    # ROW 3 — empty spacer
    # ════════════════════════════════════════════════════════
    ws.row_dimensions[3].height = 19.95

    # ════════════════════════════════════════════════════════
    # ROW 4 — Teacher name + Employee Number
    # ════════════════════════════════════════════════════════
    ws.row_dimensions[4].height = 30.0
    # B4: label
    put(4, 2, "Teacher's Name",
        font=_f(bold=True, sz=12, color=BLACK),
        fill=_fill(LABEL_BG), aln=_aln("left","center"), border=ALL)
    # C4: value
    put(4, 3, tname,
        font=_f(bold=False, sz=12, color=BLACK, name="Calibri"),
        aln=_aln("left","center"), border=ALL)
    # D4: label
    put(4, 4, "Employee Number",
        font=_f(bold=True, sz=12, color=WHITE),
        fill=_fill(EMP_BG), aln=_aln("center","center"),
        border=Border(right=_thin, top=_thin, bottom=_thin))
    # E4: value
    put(4, 5, "",
        font=_f(bold=False, sz=12, color=BLACK, name="Calibri"),
        aln=_aln("left","center"), border=ALL)

    # ════════════════════════════════════════════════════════
    # ROW 5 — Subjects Taught
    # ════════════════════════════════════════════════════════
    ws.row_dimensions[5].height = 27.0
    put(5, 2, "Subjects Taught",
        font=_f(bold=True, sz=12, color=BLACK),
        fill=_fill(LABEL_BG), aln=_aln("left","center"), border=ALL)
    put(5, 3, tsubj,
        font=_f(bold=False, sz=12, color=BLACK),
        aln=_aln("left","center"), border=ALL)

    # ════════════════════════════════════════════════════════
    # ROW 6 — Section
    # ════════════════════════════════════════════════════════
    ws.row_dimensions[6].height = 57.0
    put(6, 2,
        "Section to which the Teacher Belong to:\n"
        "[Primary/Middle School/Lower Secondary/Upper Secondary]",
        font=_f(bold=True, sz=12, color=BLACK),
        fill=_fill(LABEL_BG), aln=_aln("left","center",wrap=True), border=ALL)
    put(6, 3, school_level,
        font=_f(bold=False, sz=12, color=BLACK),
        aln=_aln("left","center",wrap=True), border=ALL)

    # ════════════════════════════════════════════════════════
    # ROW 7 — Supervising Sectional Head
    # ════════════════════════════════════════════════════════
    ws.row_dimensions[7].height = 19.95
    put(7, 2, "Supervising Sectional Head's Name",
        font=_f(bold=True, sz=12, color=BLACK),
        fill=_fill(LABEL_BG), aln=_aln("left","center"), border=ALL)
    put(7, 3, "",
        font=_f(bold=False, sz=12, color=BLACK),
        aln=_aln("left","center"), border=ALL)
    put(7, 4, "Employee Number",
        font=_f(bold=True, sz=12, color=WHITE),
        fill=_fill(EMP_BG), aln=_aln("center","center"),
        border=Border(right=_thin, top=_thin, bottom=_thin))
    put(7, 5, "",
        font=_f(bold=False, sz=12, color=BLACK),
        aln=_aln("left","center"), border=ALL)

    # ════════════════════════════════════════════════════════
    # ROWS 8-9 — spacers (no height set = default)
    # ════════════════════════════════════════════════════════

    # ════════════════════════════════════════════════════════
    # ROW 10 — Column headers  TIME / MON / TUE / WED / THU / FRI
    # ════════════════════════════════════════════════════════
    ws.row_dimensions[10].height = 19.95
    for c, txt in enumerate(["TIME","MONDAY","TUESDAY","WEDNESDAY","THURSDAY","FRIDAY"], 2):
        put(10, c, txt,
            font=_f(bold=True, sz=16, color=BLACK),
            aln=_aln("center","center"), border=ALL)

    # ════════════════════════════════════════════════════════
    # ROWS 11-21 — Time slots
    # Exact time strings and row heights from template
    # ════════════════════════════════════════════════════════
    SLOT_DATA = [
        # (sn, time_string,               label,              is_break, row, height)
        (1,  "7.40 a.m. - 8.15 a.m.\xa0",  "",                 False, 11, 19.95),
        (2,  "8.15 a.m. - 8.30 a.m. ",      "Register Marking", True,  12, 22.05),
        (3,  "8.30 a.m. - 9.10 a.m.\xa0",  "",                 False, 13, 19.05),
        (4,  "9.10 a.m. - 9.50 a.m.\xa0",  "",                 False, 14, 19.95),
        (5,  "9.50 a.m. - 10.25 a.m.\xa0", "",                 False, 15, 19.05),
        (6,  "10.25 a.m. - 10.50 a.m.",      "Interval",         True,  16, 22.95),
        (7,  "10.50 a.m. - 10.55 a.m. ",     "Seiri Time",       True,  17, 22.05),
        (8,  "10.55 a.m. - 11.35 a.m.\xa0", "",                False, 18, 19.95),
        (9,  "11.35 a.m. - 12.15 p.m.\xa0", "",                False, 19, 19.05),
        (10, "12.15 p.m. - 1.00 p.m.\xa0",  "",                False, 20, 19.05),
        (12, "1.00 p.m. - 1.45 p.m.\xa0",   "",                False, 21, 24.0),
    ]

    for sn, time_str, lbl, is_break, row, height in SLOT_DATA:
        ws.row_dimensions[row].height = height

        # B col — time string; left-align, top-align, wrap
        put(row, 2, time_str,
            font=_f(bold=False, sz=12, color=BLACK, name="Calibri"),
            aln=Alignment(horizontal="justify" if is_break else None,
                          vertical="top", wrap_text=True),
            border=ALL)

        if is_break:
            # Merge C:G → yellow background, bold black, centered
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
            ce = ws.cell(row=row, column=3, value=lbl)
            ce.font      = _f(bold=True, sz=12, color=BLACK)
            ce.fill      = _fill(BREAK_BG)
            ce.alignment = _aln("center","center")
            ce.border    = ALL
            # Merged trailing cells need top/bottom borders only (right border on G)
            for c in range(4, 7):
                ws.cell(row=row, column=c).border = TB
            ws.cell(row=row, column=7).border = RTB
        else:
            # Teaching slots — C:G each get class name or blank
            pidx = PERIOD_SNS.index(sn)
            for ci, d in enumerate(range(5), 3):
                cls_name = t_sched.get((d, pidx))
                if cls_name:
                    # Show "Grade Class" e.g. "6A" or full name
                    display = class_short.get(cls_name, cls_name)
                    put(row, ci, display,
                        font=_f(bold=True, sz=12, color=BLACK),
                        aln=_aln("center","center",wrap=True), border=ALL)
                else:
                    put(row, ci, "",
                        font=_f(bold=False, sz=12, color=BLACK),
                        aln=_aln("center","center",wrap=True), border=ALL)

    # ════════════════════════════════════════════════════════
    # ROWS 22-24 — spacers
    # ════════════════════════════════════════════════════════
    for r in [22, 23, 24]:
        ws.row_dimensions[r].height = 13.05

    # ════════════════════════════════════════════════════════
    # ROW 25 — Important notes  B25:G25
    # light blue bg, Arial Bold 12, justified, all-borders
    # ════════════════════════════════════════════════════════
    ws.row_dimensions[25].height = 106.95
    notes = (
        "IMPORTANT POINTS TO BE NOTED:\n"
        "1. This format should be used by all the schools WITHOUT FAIL. "
        "No any formats can be used to prepare Subject Timetables.\n"
        "2. The subject teacher will be officially assigned to the section in which "
        "she conducts the highest number of periods, based on her personal timetable. "
        "Accordingly, the Sectional Head of that section will serve as her designated "
        "supervising Sectional Head.\n"
        "3. Grades 1 to 5 - Either the Class teachers or the co-class/assistant teachers "
        "are expected to be in their respective classrooms during the last period "
        "(till 1:45 pm), even if no formal teaching is conducted.\n"
        "4. Each new subject teacher should be assigned a senior teacher preferably "
        "from the same subject area for guidance and observation for one term. "
        "These observation periods must be included in the newly recruited teachers\' "
        "individual timetables and highlighted in yellow for visibility."
    )
    ws.merge_cells(start_row=25, start_column=2, end_row=25, end_column=7)
    ce = ws.cell(row=25, column=2, value=notes)
    ce.font      = Font(name="Arial Bold", bold=True, size=12, color=BLACK)
    ce.fill      = _fill(NOTES_BG)
    ce.alignment = _aln("justify","center",wrap=True)
    ce.border    = ALL
    for c in range(3, 7):
        ws.cell(row=25, column=c).border = TB
    ws.cell(row=25, column=7).border = RTB

    # ════════════════════════════════════════════════════════
    # Column widths — exact from template
    # ════════════════════════════════════════════════════════
    ws.column_dimensions["A"].width = 8.22
    ws.column_dimensions["B"].width = 39.44
    ws.column_dimensions["C"].width = 30.33
    ws.column_dimensions["D"].width = 28.33
    ws.column_dimensions["E"].width = 31.11
    ws.column_dimensions["F"].width = 28.33
    ws.column_dimensions["G"].width = 26.66
    ws.column_dimensions["H"].width = 8.22


def generate_workbook(grp, excel_path, log_callback=None, **kwargs):
    def log(m): log_callback and log_callback(m)
    now_str = datetime.now().strftime("%d %b %Y  %H:%M")

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    mode = grp.get("setup_mode","legacy")

    # ── FLEXIBLE MODE ──────────────────────────────────────────
    if mode == "flexible":
        classes = grp["classes"]
        if "Setup" not in wb.sheetnames:
            wb.create_sheet("Setup",0)
        ws_setup = wb["Setup"]

        # Read subject rows from Setup sheet
        subject_rows = _read_setup_flexible(ws_setup, grp)
        if not subject_rows:
            # Fallback to data.py defaults
            subject_rows = []
            for srow in grp.get("subject_rows",[]):
                if isinstance(srow, dict):
                    subject_rows.append(srow)  # religion_block or already a dict
                else:
                    tch, subj, pc, mgs, sd, a1, room = srow

        # ── Sync religion_block and merge_block teacher names with Setup edits ──
        # Build a name-mapping from regular rows: original_name -> edited_name
        # by comparing data.py defaults vs what was read from Setup
        import copy
        default_rows = grp.get("subject_rows", [])
        default_names = {}  # index-based mapping: (subject, per_class_key) -> teacher
        for srow in default_rows:
            if isinstance(srow, (list, tuple)):
                tch0, subj0 = srow[0], srow[1]
                pc0 = srow[2] if len(srow) > 2 else {}
                key = (subj0, tuple(sorted(pc0.keys()))[:2] if isinstance(pc0,dict) else ())
                default_names[key] = tch0

        edited_names = {}  # same key -> edited teacher name
        for srow in subject_rows:
            if not isinstance(srow, dict) or srow.get("type"):
                continue
            tch0 = srow.get("teacher",""); subj0 = srow.get("subject","")
            pc0 = srow.get("per_class",{})
            key = (subj0, tuple(sorted(pc0.keys()))[:2] if isinstance(pc0,dict) else ())
            if key in default_names and default_names[key] != tch0:
                edited_names[default_names[key]] = tch0  # old -> new

        # Apply renames to religion_block and merge_block dicts
        if edited_names:
            new_rows = []
            for srow in subject_rows:
                if isinstance(srow, dict) and srow.get("type") in ("religion_block","merge_block"):
                    srow = copy.deepcopy(srow)
                    for t in srow.get("teachers", []):
                        if t["teacher"] in edited_names:
                            t["teacher"] = edited_names[t["teacher"]]
                new_rows.append(srow)
            subject_rows = new_rows
        log(f"  {len(subject_rows)} subject rows loaded from Setup")

        # Schedule
        log("  Running flexible scheduler …")
        from scheduler import schedule_flexible
        # Use shared_teacher_busy if provided (for multi-grade scheduling)
        shared_busy = kwargs.get("shared_teacher_busy", None)
        # Best-of-N: the placement RULES are identical on every attempt; only the
        # tie-break ordering is perturbed (attempt>0). On a 100%-full grid this
        # lets the scheduler explore a few layouts and keep the first that is
        # fully conflict-free — exactly how a human re-tries a tight timetable.
        max_attempts = kwargs.get("max_attempts", 80)
        base_busy = set(shared_busy) if shared_busy is not None else None
        best = None
        from scheduler import maths_layout_score, count_break_straddles
        for attempt in range(max_attempts):
            trial_busy = set(base_busy) if base_busy is not None else None
            all_slots, conflicts, double_booked = schedule_flexible(
                classes, subject_rows, shared_teacher_busy=trial_busy, attempt=attempt)
            n_conf = sum(conflicts.values())
            splits, doubles = maths_layout_score(all_slots, classes)
            straddles = count_break_straddles(all_slots, classes)
            # rank: fewest unplaced, then fewest break-straddling "broken doubles"
            # (any subject), then fewest split Maths days, then most real doubles
            score = (n_conf, straddles, splits, -doubles)
            if best is None or score < best[0]:
                best = (score, all_slots, conflicts, double_booked, trial_busy)
            if n_conf == 0 and straddles == 0 and splits == 0:
                break
        _, all_slots, conflicts, double_booked, trial_busy = best
        # Commit the winning attempt's cross-grade teacher usage back to shared_busy
        if shared_busy is not None and trial_busy is not None:
            shared_busy.clear(); shared_busy.update(trial_busy)
        if conflicts:
            log(f"  (best of {attempt+1} attempt(s): {sum(conflicts.values())} unplaced)")
        else:
            _sp, _db = maths_layout_score(all_slots, classes)
            log(f"  ✅  Conflict-free layout (attempt {attempt+1}; "
                f"maths split-days={_sp})")

        for cls in classes:
            placed = sum(1 for d in range(1,6) for p in PERIOD_SNS if all_slots[cls][d][p])
            conf   = sum(1 for (c,sk),v in conflicts.items() if c==cls and v>0)
            dist   = [sum(1 for p in PERIOD_SNS if all_slots[cls][d][p]) for d in range(1,6)]
            status = "✅" if conf==0 else f"⚠ {conf} unplaced"
            log(f"    {cls:<6} {status}  dist={dist}")

        if double_booked:
            log(f"  ⚠  {len(double_booked)} double-booking(s)")
        else:
            log("  ✅  Zero double-bookings")

        # Write class sheets
        hc = grp.get("stream_color","1F3864")
        for cls in classes:
            if cls in wb.sheetnames: del wb[cls]
            ws2 = wb.create_sheet(cls)
            n_conf = sum(1 for (c,sk),v in conflicts.items() if c==cls and v>0)
            write_class_sheet_flexible(ws2, cls, hc, all_slots[cls], now_str, n_conf)
            log(f"  Written: {cls}")

        # Teacher Overview
        sn="Teacher Overview"
        if sn in wb.sheetnames: del wb[sn]
        ws_ov=wb.create_sheet(sn)
        write_teacher_overview_flexible(ws_ov, grp, all_slots, now_str)

        # Conflict Log
        sn2="Conflict Log"
        if sn2 in wb.sheetnames: del wb[sn2]
        ws_log=wb.create_sheet(sn2)
        write_conflict_log(ws_log, conflicts, double_booked, now_str)

        # ── Personal Teacher Timetables (official Lyceum format) ──
        log("  Writing personal teacher timetables …")
        grade_num = 0
        for ch in grp["name"]:
            if ch.isdigit():
                grade_num = int(ch); break
        if grade_num <= 5:   school_level = "Primary"
        elif grade_num <= 8: school_level = "Lower Secondary"
        elif grade_num <= 11:school_level = "Upper Secondary"
        else:                school_level = "Advanced Level"

        # ── Collect all unique teachers ─────────────────────────────────────
        # Religion block: find placed slot by scanning all_slots directly.
        # This always gets the CURRENT run's slot, never a stale cached value.
        def _find_religion_slot():
            for cls in classes:
                for d in range(1, 6):
                    for p in PERIOD_SNS:
                        e = all_slots[cls][d][p]
                        if e and e[1] == "RELIGION_BLOCK":
                            return d, p
            return None, None

        rel_day, rel_per = _find_religion_slot()

        # ── Build set of all merge_block teacher names (to avoid overwriting with regular entry) ──
        mb_teacher_names = set()
        for row in subject_rows:
            if isinstance(row, dict) and row.get("type") == "merge_block":
                for t in row.get("teachers", []):
                    mb_teacher_names.add(t["teacher"])

        seen_teachers = {}

        # ── Pass 1: process block-type rows first (religion, merge) ──────────
        for row in subject_rows:
            if not isinstance(row, dict):
                continue
            rtype = row.get("type", "")

            if rtype == "religion_block":
                grade_lbl = row.get("grade_label", grp["name"])
                all_cls   = row.get("all_classes", classes)
                for t in row.get("teachers", []):
                    tch = t["teacher"]
                    if tch not in seen_teachers:
                        seen_teachers[tch] = {
                            "subject":      t["subject"],
                            "is_religion":  True,
                            "grade_label":  grade_lbl,
                            "all_classes":  all_cls,
                            "placed_day":   rel_day,
                            "placed_period":rel_per,
                            "covers":       t.get("covers", ""),
                        }

            elif rtype == "merge_block":
                for t in row.get("teachers", []):
                    tch = t["teacher"]
                    if tch not in seen_teachers:
                        seen_teachers[tch] = {
                            "subject":       t["subject"],
                            "is_religion":   False,
                            "is_merge_block":True,
                            "merge_block":   row,
                        }

        # ── Pass 2: regular rows (skip if already added as block teacher) ────
        for row in subject_rows:
            if not isinstance(row, dict):
                continue
            if row.get("type") in ("religion_block", "merge_block"):
                continue
            tch  = row.get("teacher", "")
            subj = row.get("subject", "")
            if not tch or tch == "RELIGION_BLOCK":
                continue
            if tch not in seen_teachers:
                seen_teachers[tch] = {"subject": subj, "is_religion": False,
                                      "all_subjects": [subj]}
            else:
                # Teacher teaches multiple subjects — accumulate all
                existing = seen_teachers[tch].get("all_subjects", [seen_teachers[tch]["subject"]])
                if subj not in existing:
                    existing.append(subj)
                    seen_teachers[tch]["all_subjects"] = existing
                    seen_teachers[tch]["subject"] = " / ".join(existing)

        CORE_SHEETS = {"Setup","Conflict Log","Teacher Overview"} | set(classes)
        # ── Always delete EVERY non-core sheet so personal TTs are fully rebuilt ──
        for sn3 in list(wb.sheetnames):
            if sn3 not in CORE_SHEETS:
                del wb[sn3]

        for tname, tinfo in seen_teachers.items():
            tsubj      = tinfo["subject"]
            is_rel     = tinfo.get("is_religion", False)
            sheet_name = re.sub(r'[\/:*?"<>|]', '_', tname)[:31]
            # Sheet was already deleted above — just create fresh
            ws_t = wb.create_sheet(sheet_name)

            # ── Build t_sched ─────────────────────────────────────────────────
            # Determine if this teacher has ANY regular subject rows (not religion block)
            is_pure_religion  = is_rel and not any(
                (isinstance(r, (list, tuple)) and r[0] == tname)
                or (isinstance(r, dict) and r.get("type") not in
                    ("religion_block","merge_block","aesthetic_block")
                    and r.get("teacher") == tname)
                for r in subject_rows
            )
            is_merge_block = tinfo.get("is_merge_block", False)

            t_sched = {}

            if is_pure_religion:
                # ── Pure religion teacher (RC / Hindu / Christianity / Islam) ──
                # Teaches ONLY the shared religion block slot — nothing else.
                placed_day = tinfo.get("placed_day")
                placed_per = tinfo.get("placed_period")
                covers_lbl = tinfo.get("covers", "All classes")
                if placed_day is not None and placed_per is not None:
                    pidx = PERIOD_SNS.index(placed_per)
                    t_sched[(placed_day - 1, pidx)] = covers_lbl
                log(f"    {tname:<30} [RELIGION ONLY] {covers_lbl}")

            elif is_merge_block:
                # ── Merge block teacher (Sakna/Prasad/Nimesha/Pulakshika etc.) ──
                # Build t_sched from MERGE_ markers in all_slots
                # Each pair (6A+6B) contributes `periods` slots showing the pair label
                mb   = tinfo.get("merge_block", {})
                bid  = mb.get("block_id","mb")
                pairs = mb.get("merge_pairs",[])
                for pair in pairs:
                    pair_cls = [c for c in pair if c in classes]
                    if not pair_cls: continue
                    pair_label = "+".join(sorted(pair_cls))
                    for cls in pair_cls:
                        for d in range(1,6):
                            for pidx,p in enumerate(PERIOD_SNS):
                                e = all_slots[cls][d][p]
                                if e and f"MERGE_{bid}" in str(e[1]):
                                    t_sched[(d-1, pidx)] = pair_label
                # Add standalone slots ONLY if teacher has standalone tuple rows
                # (e.g. Sakna has 6G Cookery rows → show 6G in TT)
                # (Pulakshika has no 6G rows → do NOT add individual class slots)
                has_standalone = any(
                    (isinstance(r,(list,tuple)) and len(r)>0 and r[0]==tname)
                    or (isinstance(r,dict) and r.get("type") not in ("merge_block","religion_block","aesthetic_block")
                        and r.get("teacher","")==tname)
                    for r in subject_rows
                )
                if has_standalone:
                    # Only scan classes NOT in any merge pair (standalone-only classes)
                    pair_cls_set = set(c for pair in pairs for c in pair)
                    raw_sa = {}
                    for cls in classes:
                        if cls in pair_cls_set:
                            continue  # skip merged classes — handled by pair labels above
                        for d in range(1,6):
                            for pidx,p in enumerate(PERIOD_SNS):
                                e = all_slots[cls][d][p]
                                if e and e[1] == tname:
                                    raw_sa.setdefault((d-1,pidx),[]).append(cls)
                    for key,cls_list in raw_sa.items():
                        lbl = "+".join(sorted(cls_list)) if len(cls_list)>1 else cls_list[0]
                        t_sched.setdefault(key, lbl)
                log(f"    {tname:<30} [MERGE_BLOCK:{bid}] {len(t_sched)} slots")

            else:
                # ── Regular teacher OR dual-role (Buddhism teacher) ────────────
                raw_slots = {}
                for cls in classes:
                    for d in range(1, 6):
                        for pidx, p in enumerate(PERIOD_SNS):
                            entry = all_slots[cls][d][p]
                            if entry and entry[1] == tname:
                                raw_slots.setdefault((d-1, pidx), []).append(cls)
                for key, cls_list in raw_slots.items():
                    t_sched[key] = "+".join(sorted(cls_list)) if len(cls_list)>1 else cls_list[0]

                if is_rel:
                    # Buddhism dual-role teacher: add the religion slot
                    # Show just the covers label (e.g. "6A" or "All classes")
                    placed_day = tinfo.get("placed_day")
                    placed_per = tinfo.get("placed_period")
                    covers_lbl = tinfo.get("covers", "")
                    if placed_day is not None and placed_per is not None:
                        pidx = PERIOD_SNS.index(placed_per)
                        key  = (placed_day - 1, pidx)
                        if key not in t_sched:
                            t_sched[key] = covers_lbl  # clean label e.g. "6A"
                    log(f"    {tname:<30} [DUAL:Subj+Buddhism] {len(t_sched)} periods")
                else:
                    log(f"    {tname:<30} ({len(t_sched)} periods/wk)")

            write_teacher_sheet(ws_t, "", tname, tsubj, "NAT", school_level,
                                t_sched, {c: c for c in classes}, now_str, "4472C4")

        wb.save(excel_path)
        log(f"  ✅  Saved: {os.path.basename(excel_path)}")
        if kwargs.get("return_slots"):
            return len(conflicts), len(double_booked), all_slots
        return len(conflicts), len(double_booked)

    # ── LEGACY MODE ────────────────────────────────────────────
    streams      = grp["streams"]
    CLASSES      = grp["classes"]      # dict {name: stream}
    teachers     = grp["teachers"]

    if "Setup" not in wb.sheetnames:
        wb.create_sheet("Setup",0)
    ws_setup = wb["Setup"]

    # Read teachers
    ALL_TEACHERS    = {}
    STREAM_TEACHERS = {s:[] for s in streams}
    def _read_teachers():
        if len(streams)==1:
            stream=streams[0]
            for r in range(7,7+len(teachers)+10):
                tid=str(ws_setup.cell(r,1).value or "").strip()
                if not tid: continue
                tn =str(ws_setup.cell(r,2).value or "").strip()
                ts =str(ws_setup.cell(r,3).value or "").strip()
                if tn: ALL_TEACHERS[tid]=(tn,ts,stream)
            for tid,tname,tsubj,_ in teachers:
                if stream==_: pass
            for tid,tname,tsubj,ts2 in teachers:
                STREAM_TEACHERS[ts2].append((tid,
                    ALL_TEACHERS.get(tid,(tname,tsubj,ts2))[0],
                    ALL_TEACHERS.get(tid,(tname,tsubj,ts2))[1],ts2))
        else:
            for stream,sc in [("PED",1),("NAT",8)]:
                if stream not in streams: continue
                st=[t for t in teachers if t[3]==stream]
                for ri,(_tid,_tn,_ts,_) in enumerate(st,7):
                    tid=str(ws_setup.cell(ri,sc).value or _tid).strip() or _tid
                    tn =str(ws_setup.cell(ri,sc+1).value or _tn).strip() or _tn
                    ts =str(ws_setup.cell(ri,sc+2).value or _ts).strip() or _ts
                    ALL_TEACHERS[_tid]=(tn,ts,stream)
                for tid,tname,tsubj,ts2 in [t for t in teachers if t[3]==stream]:
                    STREAM_TEACHERS[ts2].append((tid,
                        ALL_TEACHERS.get(tid,(tname,tsubj,ts2))[0],
                        ALL_TEACHERS.get(tid,(tname,tsubj,ts2))[1],ts2))
    _read_teachers()

    # Read subjects
    CLASS_SUBJECTS = _read_subjects_from_setup(ws_setup, grp)
    CLASS_SHORT = {}
    for cls in CLASSES:
        parts=cls.split()
        CLASS_SHORT[cls]=" ".join(parts[2:]) if len(parts)>=3 else (parts[-1] if parts else cls)

    # Schedule
    log("  Running constraint-based scheduler …")
    from scheduler import schedule_all
    all_slots, conflicts, double_booked = schedule_all(CLASS_SUBJECTS)

    for cls in CLASSES:
        conf=sum(1 for (c,_),v in conflicts.items() if c==cls and v>0)
        dist=[sum(1 for p in PERIOD_SNS if all_slots[cls][d][p]) for d in range(1,6)]
        status="✅" if conf==0 else f"⚠ {conf} conflict(s)"
        log(f"    {cls:<22} {status}  dist={dist}")

    if double_booked: log(f"  ⚠  {len(double_booked)} double-booking(s)")
    else: log("  ✅  Zero double-bookings")

    palette=["2E75B6","375623","843C0C","4A235A","1F6B75","7B3F00","2D6B55","5C4033"]
    CLASS_COLORS={cls:palette[i%len(palette)] for i,cls in enumerate(CLASSES)}

    for cls_name,stream in CLASSES.items():
        if cls_name not in wb.sheetnames: wb.create_sheet(cls_name)
        ws2=wb[cls_name]
        ws2.delete_rows(1,ws2.max_row+1)
        write_class_sheet(ws2,cls_name,stream,CLASS_COLORS.get(cls_name,"1F3864"),
                          all_slots,ALL_TEACHERS,CLASS_SUBJECTS,CLASS_SHORT,now_str,conflicts)
        log(f"  Written: {cls_name}")

    if "Teacher Overview" not in wb.sheetnames: wb.create_sheet("Teacher Overview")
    ws_ov=wb["Teacher Overview"]
    ws_ov.delete_rows(1,200)
    write_teacher_overview(ws_ov,grp,all_slots,ALL_TEACHERS,STREAM_TEACHERS,CLASSES)

    if "Conflict Log" not in wb.sheetnames: wb.create_sheet("Conflict Log")
    ws_log=wb["Conflict Log"]
    ws_log.delete_rows(1,200)
    write_conflict_log(ws_log,conflicts,double_booked,now_str)

    # Teacher personal sheets
    log("  Writing teacher personal timetables …")
    CORE={"Setup","Conflict Log","Teacher Overview"}|set(CLASSES.keys())
    old_names={tn[:31] for _,tn,_,_ in teachers}
    new_names={ALL_TEACHERS[tid][0][:31] for tid in ALL_TEACHERS}
    for sn in list(wb.sheetnames):
        if sn not in CORE and sn not in (old_names|new_names): del wb[sn]

    for stream in streams:
        hc=grp["stream_colors"].get(stream,"1F4E79")
        my_classes=[c for c,s in CLASSES.items() if s==stream]
        grades=set()
        for cls in my_classes:
            m=re.search(r'\b(\d{1,2})\b',cls)
            if m: grades.add(int(m.group(1)))
        mg=min(grades) if grades else 0
        sl="Lower Secondary" if mg<=8 else ("Upper Secondary" if mg<=11 else "Advanced Level")

        for tid,tname,tsubj,_ in STREAM_TEACHERS[stream]:
            uname=ALL_TEACHERS.get(tid,(tname,tsubj,""))[0]
            sn=uname[:31]
            if tname[:31] in wb.sheetnames and tname[:31]!=sn: del wb[tname[:31]]
            if sn in wb.sheetnames: del wb[sn]
            ws_t=wb.create_sheet(sn)
            t_sched={}
            for cls in my_classes:
                for d in range(1,6):
                    for si,p in enumerate(PERIOD_SNS):
                        e=all_slots[cls][d][p]
                        if e and e[2]==tid:
                            t_sched[(d-1,si)]=cls
            write_teacher_sheet(ws_t,tid,uname,tsubj,stream,sl,
                                 t_sched,CLASS_SHORT,now_str,hc)
            log(f"    {uname:<30} ({len(t_sched)} periods/wk)")

    wb.save(excel_path)
    log(f"  ✅  Saved: {os.path.basename(excel_path)}")
    if kwargs.get("return_slots"):
        return len(conflicts), len(double_booked), all_slots
    return len(conflicts), len(double_booked)


def _read_subjects_from_setup(ws, grp):
    streams=grp["streams"]
    CLASSES=grp["classes"]
    subjects_out={}

    def find_code_header(ws,col,start=10,end=80):
        for r in range(start,end):
            v=str(ws.cell(r,col).value or "").strip()
            if v=="Code": return r+1
        return None

    def read_subj_rows(ws,code_col,name_col,tid_col,pw_col,room_col,note_col,r0):
        result=[]
        if r0 is None: return result
        for r in range(r0,r0+40):
            code=str(ws.cell(r,code_col).value or "").strip()
            name=str(ws.cell(r,name_col).value or "").strip()
            if not code or not name: continue
            if code.startswith("X") and len(code)<=4: continue
            tid =str(ws.cell(r,tid_col).value or "").strip()
            try: pw=int(ws.cell(r,pw_col).value or 0)
            except: pw=0
            room=str(ws.cell(r,room_col).value or "R??").strip()
            note=str(ws.cell(r,note_col).value or "").strip()
            result.append((code,name,tid,pw,room,note))
        return result

    if len(streams)==1:
        r0=find_code_header(ws,1)
        raw=read_subj_rows(ws,1,2,3,4,5,6,r0)
        raw=[s for s in raw if s[3]>0]
        for cls in CLASSES: subjects_out[cls]=raw
    else:
        r0_ped=find_code_header(ws,1)
        r0_nat=find_code_header(ws,9)
        ped_raw=[s for s in read_subj_rows(ws,1,2,3,4,5,6,r0_ped) if s[3]>0]
        nat_raw=[s for s in read_subj_rows(ws,9,10,11,12,13,14,r0_nat) if s[3]>0]
        for cls,stream in CLASSES.items():
            if stream=="PED":
                cls_lower=cls.lower()
                is_sci="sci" in cls_lower
                is_com="com" in cls_lower
                filtered=[]
                for s in ped_raw:
                    notes=s[5].lower() if len(s)>5 else ""
                    if is_com and ("sci a&b" in notes or "sci only" in notes): continue
                    if is_sci and ("commerce only" in notes or "com only" in notes): continue
                    filtered.append(s)
                subjects_out[cls]=filtered
            else:
                subjects_out[cls]=nat_raw
    return subjects_out






# ═══════════════════════════════════════════════════════════════════════════════
def generate_all_grades(grade_groups, excel_paths, log_callback=None):
    """
    Schedule all flexible grades with shared_teacher_busy so cross-grade teachers
    (Sakna, Pulakshika, EFF Speech etc.) never double-book.

    Each grade file gets GRADE-SPECIFIC personal TTs:
      Grade 6 file → slots in Grade 6 only
      Grade 7 file → slots in Grade 7 only
      Grade 8 file → slots in Grade 8 only

    The combined multi-grade TT (all 3 grades in one sheet) is generated
    separately via export_teacher_combined_pdf() for the named teacher.
    Returns list of (conflicts, dbl) per grade AND dict of all_slots_by_grade.
    """
    from openpyxl import load_workbook, Workbook
    from scheduler import schedule_flexible, PERIOD_SNS, DAYS
    import re as _re, os
    from datetime import datetime
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    def log(m):
        if log_callback: log_callback(m)

    shared_busy    = set()
    grade_data     = []   # (grp, all_slots, subject_rows, classes, path)
    results        = []

    # ── Pass 1: schedule all grades with shared busy set ─────────────────────
    for grp, path in zip(grade_groups, excel_paths):
        if grp.get("setup_mode") != "flexible":
            continue
        if not os.path.exists(path):
            wb0 = Workbook(); wb0.remove(wb0.active); wb0.create_sheet("Setup")
            write_setup(wb0["Setup"], grp); wb0.save(path)

        result = generate_workbook(
            grp, path, log_callback=log_callback,
            shared_teacher_busy=shared_busy,
            return_slots=True,
        )
        if isinstance(result, tuple) and len(result) == 3:
            nc, ndb, all_slots = result
        else:
            nc, ndb = result[:2]; all_slots = None

        results.append((nc, ndb))
        _wb = load_workbook(path)
        subject_rows = _read_setup_flexible(_wb["Setup"], grp)
        grade_data.append((grp, all_slots, subject_rows, grp["classes"], path))

    # ── Pass 2: write GRADE-SPECIFIC personal TT sheets ──────────────────────
    for grp, all_slots, subject_rows, classes, path in grade_data:
        if all_slots is None:
            continue
        gnum = grp["name"].split()[-1]
        wb3  = load_workbook(path)
        CORE = {"Setup", "Conflict Log", "Teacher Overview"} | set(classes)

        # Delete old personal TT sheets
        for sn3 in list(wb3.sheetnames):
            if sn3 not in CORE:
                del wb3[sn3]

        # Build t_sched for each teacher using THIS grade's slots only
        teacher_sched   = {}  # teacher -> t_sched (grade-specific)
        teacher_subject = {}  # teacher -> subject label

        for row in subject_rows:
            if isinstance(row, dict):
                rtype = row.get("type","")

                if rtype == "merge_block":
                    bid   = row["block_id"]
                    pairs = row.get("merge_pairs",[])
                    for tinfo in row.get("teachers",[]):
                        tname = tinfo["teacher"]
                        if tname not in teacher_sched:
                            teacher_sched[tname]   = {}
                            teacher_subject[tname] = tinfo["subject"]
                        for pair in pairs:
                            pair_cls = [c for c in pair if c in classes]
                            if not pair_cls: continue
                            pair_lbl = "+".join(sorted(pair_cls))
                            # Write ALL slots for this pair (both periods show the label)
                            for cls in pair_cls[:1]:  # one class representative
                                for d in range(1,6):
                                    for pidx,p in enumerate(PERIOD_SNS):
                                        e = all_slots[cls][d][p]
                                        if e and f"MERGE_{bid}" in str(e[1]):
                                            teacher_sched[tname][(d-1,pidx)] = pair_lbl

                elif rtype == "religion_block":
                    rel_key = None
                    for cls in classes:
                        for d in range(1,6):
                            for pidx,p in enumerate(PERIOD_SNS):
                                e = all_slots[cls][d][p]
                                if e and e[1]=="RELIGION_BLOCK":
                                    rel_key=(d-1,pidx); break
                            if rel_key: break
                        if rel_key: break

                    for tinfo in row.get("teachers",[]):
                        tname  = tinfo["teacher"]
                        covers = tinfo.get("covers","")
                        if tname not in teacher_sched:
                            teacher_sched[tname]   = {}
                            teacher_subject[tname] = tinfo["subject"]
                        if rel_key:
                            # pure religion = only this 1 slot
                            # Buddhism (dual-role) = this slot + regular subject slots
                            teacher_sched[tname][rel_key] = covers

                if rtype in ("merge_block","religion_block"):
                    continue

            # Regular dict row (normal subject)
            tname = row.get("teacher","") if isinstance(row,dict) else (row[0] if isinstance(row,(list,tuple)) else "")
            subj  = row.get("subject","") if isinstance(row,dict) else (row[1] if isinstance(row,(list,tuple)) else "")
            if not tname: continue

            if tname not in teacher_sched:
                teacher_sched[tname]   = {}
                teacher_subject[tname] = subj
            else:
                exist = teacher_subject.get(tname,"")
                if subj not in exist:
                    teacher_subject[tname] = exist + " / " + subj

            raw = {}
            for cls in classes:
                for d in range(1,6):
                    for pidx,p in enumerate(PERIOD_SNS):
                        e = all_slots[cls][d][p]
                        if e and e[1]==tname:
                            raw.setdefault((d-1,pidx),[]).append(cls)
            for key,cls_list in raw.items():
                lbl = "+".join(sorted(cls_list)) if len(cls_list)>1 else cls_list[0]
                teacher_sched[tname][key] = lbl

        # Write one sheet per teacher with grade-specific slots
        written = 0
        for tname, t_sched in teacher_sched.items():
            if not t_sched: continue
            sn = _re.sub(r'[\\/:*?"<>|]',"_",tname)[:31]
            ws_t = wb3.create_sheet(sn)
            write_teacher_sheet(ws_t,"",tname,teacher_subject.get(tname,""),
                                "NAT","Lower Secondary",t_sched,{},now_str,"4472C4")
            written += 1

        wb3.save(path)
        log(f"  ✅ {os.path.basename(path)}: {written} teacher sheets (Grade {gnum} only)")

    # Store all_slots_by_grade in a module-level cache for PDF export
    _grade_slots_cache.clear()
    for grp, all_slots, subject_rows, classes, path in grade_data:
        _grade_slots_cache[grp["name"]] = {
            "all_slots": all_slots,
            "subject_rows": subject_rows,
            "classes": classes,
            "grp": grp,
        }

    return results

# ── Cache for multi-grade combined PDF export ─────────────────────────────────
_grade_slots_cache = {}   # grade_name -> {all_slots, subject_rows, classes, grp}


def get_teacher_combined_sched(teacher_name):
    """
    Build combined t_sched for a teacher across all scheduled grades.
    Returns dict: {(day0, pidx): "6:6A+6B | 7:7A+7B | 8:8A+8B"} for display.
    """
    from scheduler import PERIOD_SNS
    combined = {}  # (day0,pidx) -> "6:label | 7:label | 8:label"
    subject  = ""

    for grade_name, cache in _grade_slots_cache.items():
        all_slots    = cache["all_slots"]
        subject_rows = cache["subject_rows"]
        classes      = cache["classes"]
        gnum         = grade_name.split()[-1]

        for row in subject_rows:
            if isinstance(row,dict):
                rtype = row.get("type","")
                if rtype=="merge_block":
                    bid   = row["block_id"]
                    pairs = row.get("merge_pairs",[])
                    if not any(t["teacher"]==teacher_name for t in row.get("teachers",[])):
                        continue
                    if not subject:
                        subject = next(t["subject"] for t in row["teachers"] if t["teacher"]==teacher_name)
                    for pair in pairs:
                        pair_cls=[c for c in pair if c in classes]
                        if not pair_cls: continue
                        lbl=gnum+":"+"+".join(sorted(pair_cls))
                        for cls in pair_cls[:1]:  # representative class
                            for d in range(1,6):
                                for pidx,p in enumerate(PERIOD_SNS):
                                    e=all_slots[cls][d][p]
                                    if e and f"MERGE_{bid}" in str(e[1]):
                                        key=(d-1,pidx)
                                        old=combined.get(key,"")
                                        combined[key]=(old+" | "+lbl) if old else lbl
                elif rtype=="religion_block":
                    tinfo=next((t for t in row.get("teachers",[]) if t["teacher"]==teacher_name),None)
                    if not tinfo: continue
                    if not subject: subject=tinfo["subject"]
                    for cls in classes:
                        for d in range(1,6):
                            for pidx,p in enumerate(PERIOD_SNS):
                                e=all_slots[cls][d][p]
                                if e and e[1]=="RELIGION_BLOCK":
                                    key=(d-1,pidx)
                                    lbl=gnum+":"+tinfo.get("covers","")
                                    old=combined.get(key,"")
                                    combined[key]=(old+" | "+lbl) if old else lbl
                                    break
                            if key in combined: break
                        if key in combined: break
                continue

            tname = row.get("teacher","") if isinstance(row,dict) else (row[0] if isinstance(row,(list,tuple)) else "")
            if tname!=teacher_name: continue
            ssubj = row.get("subject","") if isinstance(row,dict) else (row[1] if isinstance(row,(list,tuple)) else "")
            if not subject: subject=ssubj

            for cls in classes:
                for d in range(1,6):
                    for pidx,p in enumerate(PERIOD_SNS):
                        e=all_slots[cls][d][p]
                        if e and e[1]==teacher_name:
                            key=(d-1,pidx)
                            lbl=gnum+":"+cls
                            old=combined.get(key,"")
                            combined[key]=(old+" | "+lbl) if old else lbl

    return combined, subject


def export_teacher_combined_xlsx(teacher_name, out_path):
    """
    Build a single .xlsx with one sheet per grade (Grade 6 / Grade 7 / Grade 8)
    for a teacher who teaches across multiple grades.
    Uses _grade_slots_cache populated after generate_all in web_app.py.

    Strategy: scan all_slots directly for this teacher — no subject_rows parsing
    needed.  This is robust regardless of how rows are stored (tuple vs dict).
    The teacher_name here is the sheet-safe name from the dropdown (same as
    re.sub special chars → '_'), which matches what the scheduler stores in
    slots because web_app collect_teacher_names + scheduler both use the raw
    name string from data.py.
    """
    from openpyxl import Workbook
    from scheduler import PERIOD_SNS
    from datetime import datetime
    import re as _re

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    GRADE_COLORS = {"6": "1A3660", "7": "7C3AED", "8": "0F7D59"}

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    any_written = False

    for grade_name in sorted(_grade_slots_cache.keys()):
        cache     = _grade_slots_cache[grade_name]
        all_slots = cache["all_slots"]
        classes   = cache["classes"]
        subject_rows = cache.get("subject_rows", [])
        gnum      = grade_name.split()[-1]
        hc        = GRADE_COLORS.get(gnum, "1A3660")

        t_sched = {}   # (day0, pidx) -> label
        t_subj  = ""

        # ── 1. Scan all_slots directly for this teacher (regular + merge slots) ──
        for cls in classes:
            for d in range(1, 6):
                for pidx, p in enumerate(PERIOD_SNS):
                    e = all_slots[cls][d][p]
                    if not e:
                        continue
                    slot_teacher = str(e[1]) if len(e) > 1 else ""
                    slot_subj    = str(e[0]) if len(e) > 0 else ""

                    if slot_teacher == teacher_name:
                        # Regular subject slot
                        key = (d - 1, pidx)
                        existing = t_sched.get(key, "")
                        if cls not in existing:
                            t_sched[key] = (existing + "+" + cls) if existing else cls
                        if not t_subj and slot_subj:
                            t_subj = slot_subj

        # ── 2. Handle RELIGION_BLOCK — find slot + look up which class this
        #       teacher covers from subject_rows ──────────────────────────────
        rel_key   = None
        rel_cover = ""
        rel_subj  = ""

        for cls in classes:
            for d in range(1, 6):
                for pidx, p in enumerate(PERIOD_SNS):
                    e = all_slots[cls][d][p]
                    if e and e[1] == "RELIGION_BLOCK":
                        rel_key = (d - 1, pidx)
                        break
                if rel_key:
                    break
            if rel_key:
                break

        if rel_key:
            for row in subject_rows:
                if not isinstance(row, dict) or row.get("type") != "religion_block":
                    continue
                for tinfo in row.get("teachers", []):
                    if tinfo.get("teacher") == teacher_name:
                        rel_cover = tinfo.get("covers", "")
                        rel_subj  = tinfo.get("subject", "")
                        break
                if rel_cover:
                    break

            if rel_cover:
                # Only add religion slot if not already captured above
                if rel_key not in t_sched:
                    t_sched[rel_key] = rel_cover
                if not t_subj and rel_subj:
                    t_subj = rel_subj

        if not t_sched:
            continue  # teacher has no slots in this grade

        sn   = f"Grade {gnum}"
        ws_t = wb_out.create_sheet(sn)
        write_teacher_sheet(
            ws_t, "", teacher_name, t_subj or "—",
            "NAT", "Lower Secondary",
            t_sched, {}, now_str, hc
        )
        any_written = True

    if any_written:
        wb_out.save(out_path)
    return any_written
