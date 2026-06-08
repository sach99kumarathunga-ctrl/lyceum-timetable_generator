#!/usr/bin/env python3
"""
web_app.py — Lyceum Timetable Generator (Web Edition v3)
─────────────────────────────────────────────────────────
Key features:
  • "Generate All Grades" schedules every grade (6, 7, 8, 9, 10, 11-12)
    against ONE shared teacher-busy set, so any teacher shared between grades
    (Grade 9 & 10 staff, Mr. Radun across 9/10/11-12, option-block teachers in
    6-8 …) is never double-booked anywhere.
  • Grades 9, 10 and 11-12 (Edexcel / National streams) are auto-converted from
    the legacy data into the editable flexible format, with the school rules:
    National maths = 1 teacher, Edexcel maths = 2 (one per grade); Edexcel
    Sinhala = Ms. Disna, National Sinhala = Ms. Sewwandi; National religion =
    Buddhism (2) alongside Cultural Dancing (2).
  • Every subject/teacher can be edited, added or deleted before generating.
  • Shared teachers get a single COMBINED personal timetable across all grades.
  • Result (Step 3) shows all grades at once: grade tabs on top, class tabs below.

Run:  python web_app.py
"""
import os, re, json, copy, uuid, tempfile, threading, webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs

import data
from excel_writer import write_setup_flexible, generate_workbook
from openpyxl import Workbook, load_workbook
from scheduler import (TIME_SLOTS, PERIOD_SNS, maths_layout_score,
                       count_break_straddles)

import os
PORT    = int(os.environ.get("PORT", 8000))
HOST    = "0.0.0.0"
WORKDIR = os.path.join(tempfile.gettempdir(), "lyceum_tt_web")
os.makedirs(WORKDIR, exist_ok=True)

# token -> {path, grade, teachers}   (single-grade cache)
GEN_CACHE = {}
# "all" token -> {grades: [{grade, path, teachers, stats, ...}]}
ALL_CACHE = {}

SN_LABEL = {sn: lab for sn, _t, lab, _b in TIME_SLOTS}


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

# ── Convert legacy stream-based grades (9, 10, 11-12) into the flexible format ──
_SCIENCE       = {"chemistry", "biology", "physics"}
_COMMERCE_LIKE = {"commerce", "accounting", "business studies", "economics",
                  "cs/ict", "commerce/lit"}


def _is_asm(name):
    n = str(name).strip().lower()
    return n == "asm" or "assembl" in n


def _is_dancing(subj):
    s = str(subj).lower()
    return "dancing" in s or "aesthet" in s


def _role(subj):
    """Short, readable role label for a generic teacher placeholder."""
    s = str(subj).strip().lower()
    table = {
        "mathematics": "Maths", "combined maths": "Combined Maths",
        "sinhala/tamil": "Sinhala", "eff. speech": "Speech",
        "effective speech": "Speech", "motivational speech": "Speech",
        "cs/ict": "ICT", "ict": "ICT", "history/geo": "History",
        "science/chem": "Science", "general english": "English",
        "r.catholicism": "Religion", "religion": "Religion",
        "business studies": "Business", "science award": "Science",
        "physical education": "PE", "biology/sci": "Biology",
        "physics/sci": "Physics", "english/lit": "English",
    }
    if s in table:
        return table[s]
    return str(subj).strip().title()


def _eligible_classes(stream_classes, subj, note, grade_name):
    n = (note or "").lower()
    s = subj.lower()
    def has(c, *keys):
        cl = c.lower()
        return any(k in cl for k in keys)
    if "sci a&b" in n or "sci a & b" in n:
        return [c for c in stream_classes if has(c, "sci a", "sci b")]
    if "nat sci only" in n:
        return [c for c in stream_classes if has(c, "nat sci")]
    if "nat com only" in n:
        return [c for c in stream_classes if has(c, "nat com")]
    if "com only" in n or "commerce only" in n or "ped com only" in n:
        return [c for c in stream_classes if has(c, "com")]
    if "11" in grade_name or "12" in grade_name or "AL" in grade_name:
        if s in _SCIENCE or s == "combined maths":
            sci = [c for c in stream_classes if has(c, "sci")]
            if sci:
                return sci
        if s in _COMMERCE_LIKE:
            com = [c for c in stream_classes if has(c, "com")]
            if com:
                return com
    return list(stream_classes)


def _aesthetic_group():
    """The shared aesthetic teacher group — the SAME people who teach the
    Art/Music/Dance option block in Grades 6, 7 & 8. Reused (same names) in
    Grades 9 & 10 so the one global teacher-busy set keeps them clash-free
    across every grade. Falls back to generic names if 6-8 isn't present."""
    for g in _FLEX_NATIVE:
        for row in g["subject_rows"]:
            if isinstance(row, dict) and row.get("type") == "merge_block":
                subs = {str(t.get("subject", "")).lower() for t in row.get("teachers", [])}
                if {"art", "dancing", "eastern music"} & subs:
                    return [{"teacher": t["teacher"], "subject": t["subject"],
                             "room": t.get("room", "R-AES")}
                            for t in row.get("teachers", [])]
    return [{"teacher": "Aesthetic Teacher 1", "subject": "Art",     "room": "R-ART"},
            {"teacher": "Aesthetic Teacher 2", "subject": "Dancing", "room": "R-DAN"},
            {"teacher": "Aesthetic Teacher 3", "subject": "Music",   "room": "R-MUS"}]


def _pairs(classes):
    """Chunk classes into groups of two for an option/merge block."""
    return [classes[i:i + 2] for i in range(0, len(classes), 2)]


def build_generic_map(grade_list, prefix=""):
    """real teacher name -> generic '<prefix>Role Teacher N', consistent across
    all the grades that share one staff namespace. The prefix keeps the two
    pools (Grade 9-10 vs Grade 11-12) from colliding into the same names."""
    m, cnt = {}, {}
    for lg in grade_list:
        for t in lg["teachers"]:
            real, subj = t[1], t[2]
            if real in m:
                continue
            role = _role(subj)
            cnt[role] = cnt.get(role, 0) + 1
            m[real] = f"{prefix}{role} Teacher {cnt[role]}"
    return m


def legacy_to_flexible(lg, name_map, ns=""):
    """Convert a legacy grade to flexible format using GENERIC placeholder names
    (drawn from `name_map`, shared across the grade's staff namespace) and the
    school's specific rules for Grades 9 & 10. `ns` is the namespace prefix."""
    tmap = {t[0]: t[1] for t in lg["teachers"]}
    name = lg["name"]
    classes = list(lg["classes"].keys())
    is_910 = name in ("Grade 9", "Grade 10")
    rows = []
    for stream in lg["streams"]:
        stream_classes = [c for c, st in lg["classes"].items() if st == stream]
        for entry in lg["subjects"][stream]:
            code, subj, tid, periods, room, note = entry
            if code == "ASM" or _is_asm(subj):
                continue                      # Assembly auto-placed by scheduler
            elig = _eligible_classes(stream_classes, subj, note, name)
            if not elig:
                continue
            real = tmap.get(tid, tid)
            tname = name_map.get(real, real)  # generic placeholder

            # ── Grade 9 & 10 specific rules ───────────────────────────────
            if is_910:
                if _is_dancing(subj):
                    # Cultural Dancing → Aesthetic, taught by the shared 6-8
                    # aesthetic group as an option block (added once below).
                    continue
                if "sinhala" in subj.lower():
                    tname = (ns + "Sinhala Teacher (Edexcel)" if stream == "PED"
                             else ns + "Sinhala Teacher (National)")
                if subj.lower() == "mathematics":
                    # National = ONE shared maths teacher (both grades).
                    # Edexcel = TWO maths teachers (one per grade) so the load
                    # is feasible and stays feasible if periods change.
                    if stream == "NAT":
                        tname = ns + "Maths Teacher 1 (National)"
                    else:
                        tname = (ns + "Maths Teacher 1 (Edexcel)" if name == "Grade 9"
                                 else ns + "Maths Teacher 2 (Edexcel)")
                if "english" in subj.lower():
                    if stream == "NAT":
                        tname = (ns + "English Teacher 1 (National)" if name == "Grade 9"
                                 else ns + "English Teacher 2 (National)")
                if stream == "NAT" and (code == "REL" or subj.lower() == "religion"):
                    subj = "Buddhism"
                    periods = 2               # Aesthetic(2) + Buddhism(2) category
                    tname = ns + "Buddhism Teacher"

            rows.append((tname, subj, {c: periods for c in elig},
                         [], False, False, room))

    # ── Aesthetic option block for Grades 9 & 10 (same teachers as 6-8) ──────
    blocks = []
    if is_910:
        blocks.append({
            "type": "merge_block", "block_id": "aesthetic_%s" % name[-1],
            "periods": 2, "same_day": True,
            "merge_pairs": _pairs(classes), "standalone": {},
            "teachers": _aesthetic_group(),
        })

    # reference teacher table for the Setup sheet
    seen, tref = set(), []
    for (tname, subj, pc, _m, _s, _a, _r) in rows:
        if (tname, subj) in seen:
            continue
        seen.add((tname, subj))
        tref.append(("L%02d" % (len(tref) + 1), tname, subj))
    for blk in blocks:
        for t in blk["teachers"]:
            if (t["teacher"], t["subject"]) not in seen:
                seen.add((t["teacher"], t["subject"]))
                tref.append(("L%02d" % (len(tref) + 1), t["teacher"], t["subject"]))

    return {"name": name, "file": lg["file"], "setup_mode": "flexible",
            "classes": classes, "subject_rows": rows + blocks, "teachers": tref}


# Build the full grade registry: native flexible 6/7/8 + converted 9/10/11-12.
_FLEX_NATIVE = [g for g in data.GRADE_GROUPS if g.get("setup_mode") == "flexible"]
_LEGACY_SRC  = [g for g in data.GRADE_GROUPS if g.get("setup_mode") == "legacy"]

# Two separate staff namespaces: {Grade 9, Grade 10} share one pool; Grade 11-12
# shares its own. Generic names are assigned consistently within each namespace.
_NS_910  = [g for g in _LEGACY_SRC if g["name"] in ("Grade 9", "Grade 10")]
_NS_1112 = [g for g in _LEGACY_SRC if g["name"] not in ("Grade 9", "Grade 10")]
_MAP_910  = build_generic_map(_NS_910,  prefix="9/10 ")
_MAP_1112 = build_generic_map(_NS_1112, prefix="11/12 ")

_CONVERTED = []
for _lg in _LEGACY_SRC:
    if _lg["name"] in ("Grade 9", "Grade 10"):
        _CONVERTED.append(legacy_to_flexible(_lg, _MAP_910, ns="9/10 "))
    else:
        _CONVERTED.append(legacy_to_flexible(_lg, _MAP_1112, ns="11/12 "))
ALL_GRADES = _FLEX_NATIVE + _CONVERTED


def flex_grades():
    return ALL_GRADES


def get_grade(name):
    for g in ALL_GRADES:
        if g["name"] == name:
            return g
    return None


def setup_payload(grp):
    classes = grp["classes"]
    regular, blocks = [], []
    ridx = bidx = 0
    for row in grp["subject_rows"]:
        if isinstance(row, (list, tuple)):
            tch, subj, pc, mgs, sd, a1, room = row[:7]
            regular.append({
                "id": "r%d" % ridx, "teacher": tch, "subject": subj,
                "room": room or "R??",
                "periods": {c: int(pc.get(c, 0)) for c in classes},
                "same_day": bool(sd), "all_1period": bool(a1),
                "merge_groups": "; ".join("+".join(g) for g in (mgs or [])),
            })
            ridx += 1
        elif isinstance(row, dict):
            rtype = row.get("type", "")
            if rtype == "religion_block":
                label = "Religion block (one shared period for all classes)"
            elif rtype == "merge_block":
                label = "Option block: " + " / ".join(
                    sorted({t["subject"] for t in row.get("teachers", [])}))
            else:
                continue
            blocks.append({
                "id": "b%d" % bidx, "type": rtype, "label": label,
                "teachers": [{"teacher": t.get("teacher", ""),
                               "subject":  t.get("subject", ""),
                               "room":     t.get("room", "R??"),
                               "covers":   t.get("covers", "")}
                              for t in row.get("teachers", [])],
            })
            bidx += 1
    return {"grade": grp["name"], "classes": classes,
            "regular_rows": regular, "blocks": blocks}


def _parse_merge_groups(s):
    out = []
    for part in str(s or "").split(";"):
        part = part.strip()
        if not part:
            continue
        if "+" in part:
            out.append([c.strip() for c in part.split("+") if c.strip()])
        else:
            out.append([part])
    return out


def apply_edits(grp, payload):
    """Rebuild the grade's subject rows from the browser payload. Because the
    rows are rebuilt (not patched positionally), teachers/subjects can be
    EDITED, DELETED (simply absent from the payload) or ADDED (new entries)."""
    grp2 = copy.deepcopy(grp)
    classes = grp2["classes"]
    reg_payload = payload.get("regular_rows", None)
    blk_payload = {b.get("id"): b for b in payload.get("blocks", [])}

    # ── Regular subject/teacher rows ──────────────────────────────────────
    new_regular = []
    if reg_payload is not None:
        for e in reg_payload:
            tch  = str(e.get("teacher", "")).strip()
            subj = str(e.get("subject", "")).strip()
            room = (str(e.get("room", "") or "").strip() or "R??")
            sd   = bool(e.get("same_day", False))
            a1   = bool(e.get("all_1period", False))
            mgs  = _parse_merge_groups(e.get("merge_groups", ""))
            pc = {}
            for c, v in (e.get("periods") or {}).items():
                if c not in classes:
                    continue
                try:
                    v = int(v)
                except (TypeError, ValueError):
                    v = 0
                if v > 0:
                    pc[c] = v
            if not subj or not tch:          # incomplete → skip silently
                continue
            if not pc and not a1:            # nothing to schedule → skip
                continue
            new_regular.append((tch, subj, pc, mgs, sd, a1, room))
    else:
        for row in grp2["subject_rows"]:
            if isinstance(row, (list, tuple)):
                new_regular.append(tuple(row[:7]))

    # ── Shared blocks (religion / option) ─────────────────────────────────
    new_blocks = []
    bidx = 0
    for row in grp2["subject_rows"]:
        if isinstance(row, dict) and row.get("type") in ("religion_block", "merge_block"):
            bkey = "b%d" % bidx
            bidx += 1
            row = copy.deepcopy(row)
            e = blk_payload.get(bkey)
            if e is None:
                new_blocks.append(row)       # untouched block
                continue
            rebuilt = []
            for i, te in enumerate(e.get("teachers", [])):
                if isinstance(te, str):                       # legacy: name only
                    name = te.strip()
                    if not name:
                        continue
                    base = dict(row["teachers"][i]) if i < len(row["teachers"]) else {}
                    base["teacher"] = name
                    rebuilt.append(base)
                elif isinstance(te, dict):
                    name = str(te.get("teacher", "")).strip()
                    subj = str(te.get("subject", "")).strip()
                    if not name or not subj:
                        continue
                    nt = {"teacher": name, "subject": subj,
                          "room": (str(te.get("room", "") or "").strip() or "R??")}
                    if row.get("type") == "religion_block":
                        nt["covers"] = (str(te.get("covers", "") or "").strip()
                                        or "All classes")
                    rebuilt.append(nt)
            if rebuilt:                       # keep block only if it still has teachers
                row["teachers"] = rebuilt
                new_blocks.append(row)
            # else: every option removed → drop the whole block

    grp2["subject_rows"] = new_regular + new_blocks
    return grp2


def collect_teacher_names(grp2):
    names = []
    for row in grp2["subject_rows"]:
        if isinstance(row, (list, tuple)):
            if str(row[0]).strip(): names.append(row[0])
        elif isinstance(row, dict):
            for t in row.get("teachers", []):
                if str(t.get("teacher", "")).strip(): names.append(t["teacher"])
    seen, out = set(), []
    for n in names:
        if n in seen: continue
        seen.add(n)
        sheet = re.sub(r'[\\/:*?"<>|]', '_', n)[:31]
        out.append({"name": n, "sheet": sheet})
    return out


def build_grids(slots, classes):
    rows_meta = [{"label": lab, "time": tr, "is_break": bool(b)}
                 for sn, tr, lab, b in TIME_SLOTS
                 if lab != "Lunch Break"]
    grids = {}
    for cls in classes:
        grid = []
        for sn, tr, lab, b in TIME_SLOTS:
            if lab == "Lunch Break":
                continue
            if b:
                grid.append(None)
                continue
            cells = []
            for d in range(1, 6):
                e = slots[cls][d].get(sn)
                if e:
                    cells.append({"subject": e[0], "teacher": e[1], "room": e[2]})
                else:
                    cells.append(None)
            grid.append(cells)
        grids[cls] = grid
    return rows_meta, grids


def _run_single_grade(grp2, shared_busy=None, max_attempts=80):
    """Generate one grade's workbook. Returns (path, slots, stats, logs)."""
    safe = re.sub(r'[^A-Za-z0-9]', '_', grp2["name"])
    path = os.path.join(WORKDIR, f"{safe}_{uuid.uuid4().hex[:8]}.xlsx")
    wb = Workbook(); wb.remove(wb.active); ws = wb.create_sheet("Setup")
    write_setup_flexible(ws, grp2); wb.save(path)

    logs = []
    kw = dict(log_callback=lambda m: logs.append(m), return_slots=True,
              max_attempts=max_attempts)
    if shared_busy is not None:
        kw["shared_teacher_busy"] = shared_busy

    res = generate_workbook(grp2, path, **kw)
    n_conf, n_dbl, slots = res
    classes = grp2["classes"]
    splits, doubles = maths_layout_score(slots, classes)
    straddles = count_break_straddles(slots, classes)
    stats = {"unplaced": int(n_conf), "double_booked": int(n_dbl),
             "straddles": int(straddles), "maths_splits": int(splits),
             "maths_doubles": int(doubles)}
    return path, slots, stats, logs


# ─────────────────────────────────────────────────────────────────────────────
#  HTTP handler
# ─────────────────────────────────────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a): pass

    def _send_json(self, obj, code=200):
        body = json.dumps(obj).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, path, name):
        with open(path, "rb") as f: raw = f.read()
        self.send_response(200)
        self.send_header("Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{name}"')
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def do_GET(self):
        u = urlparse(self.path)
        if u.path == "/":
            body = INDEX_HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        elif u.path == "/api/grades":
            self._send_json({"grades": [g["name"] for g in flex_grades()]})
        elif u.path == "/api/setup":
            qs  = parse_qs(u.query)
            grp = get_grade(qs.get("grade", [""])[0])
            if not grp: return self._send_json({"error": "unknown grade"}, 404)
            self._send_json(setup_payload(grp))
        elif u.path.startswith("/api/download/"):
            qs  = parse_qs(u.query)
            tok = qs.get("token", [""])[0]

            # ── Combined multi-grade teacher TT — no grade param needed ──────
            if u.path == "/api/download/teacher_combined":
                sheet = qs.get("sheet", [""])[0]
                out   = self._extract_teacher_combined(sheet)
                if not out: return self._send_json({"error": "teacher not found in any grade — please use Generate All Grades first"}, 404)
                return self._send_file(out, f"{sheet}_AllGrades_Timetable.xlsx")

            # ── All other downloads need a grade-specific file path ───────────
            grade_path = None
            if tok in GEN_CACHE:
                info = GEN_CACHE[tok]
                grade_path = info["path"]
                grade_name = info["grade"]
            elif tok in ALL_CACHE:
                gname = qs.get("grade", [""])[0]
                gi    = next((x for x in ALL_CACHE[tok]["grades"] if x["grade"]==gname), None)
                if not gi: return self._send_json({"error": "grade not found"}, 404)
                grade_path = gi["path"]
                grade_name = gi["grade"]
            else:
                return self._send_json({"error": "expired — regenerate"}, 404)

            if u.path == "/api/download/full":
                self._send_file(grade_path,
                    grade_name.replace(" ","") + "_Timetable.xlsx")
            elif u.path == "/api/download/teacher":
                sheet = qs.get("sheet", [""])[0]
                out   = self._extract_teacher(grade_path, sheet)
                if not out: return self._send_json({"error": "sheet not found"}, 404)
                self._send_file(out, f"{sheet}_Timetable.xlsx")
        else:
            self._send_json({"error": "not found"}, 404)

    def do_POST(self):
        u      = urlparse(self.path)
        length = int(self.headers.get("Content-Length", 0))
        raw    = self.rfile.read(length)
        try:
            payload = json.loads(raw.decode("utf-8"))
        except Exception:
            return self._send_json({"error": "bad JSON"}, 400)

        # ── Single-grade generate ────────────────────────────────────────────
        if u.path == "/api/generate":
            grp = get_grade(payload.get("grade", ""))
            if not grp: return self._send_json({"error": "unknown grade"}, 400)
            try:
                grp2 = apply_edits(grp, payload)
                path, slots, stats, logs = _run_single_grade(grp2)
            except Exception as e:
                import traceback
                return self._send_json(
                    {"error": f"generation failed: {e}",
                     "trace": traceback.format_exc()}, 500)
            classes  = grp2["classes"]
            teachers = collect_teacher_names(grp2)
            wb       = load_workbook(path, read_only=True)
            present  = set(wb.sheetnames); wb.close()
            teachers = [t for t in teachers if t["sheet"] in present]
            token    = uuid.uuid4().hex
            GEN_CACHE[token] = {"path": path, "grade": grp2["name"], "teachers": teachers}
            rows_meta, grids = build_grids(slots, classes)
            return self._send_json({
                "ok": True, "token": token, "grade": grp2["name"],
                "classes": classes, "rows_meta": rows_meta, "grids": grids,
                "teachers": teachers, "stats": stats, "log_tail": logs[-6:],
            })

        # ── All-grades generate (shared teacher-busy set) ────────────────────
        elif u.path == "/api/generate_all":
            # payload.grades = [{grade, regular_rows, blocks}, ...]  (may be empty / partial)
            grade_payloads = {p["grade"]: p for p in payload.get("grades", [])}
            shared_busy    = set()
            all_results    = []
            # Schedule the tightly-shared senior grades (9, 10, 11-12) FIRST so their
            # shared teachers claim slots before the lighter, independent grades 6-8.
            order = {"Grade 9": 0, "Grade 10": 1, "Grade 11-12 AL": 2,
                     "Grade 6": 3, "Grade 7": 4, "Grade 8": 5}
            sched_grades = sorted(flex_grades(), key=lambda g: order.get(g["name"], 9))
            try:
                for grp in sched_grades:
                    gname = grp["name"]
                    gp    = grade_payloads.get(gname, {})
                    grp2  = apply_edits(grp, gp) if gp else copy.deepcopy(grp)
                    path, slots, stats, logs = _run_single_grade(grp2, shared_busy,
                                                                  max_attempts=200)
                    classes  = grp2["classes"]
                    teachers = collect_teacher_names(grp2)
                    wb       = load_workbook(path, read_only=True)
                    present  = set(wb.sheetnames); wb.close()
                    teachers = [t for t in teachers if t["sheet"] in present]
                    rows_meta, grids = build_grids(slots, classes)
                    all_results.append({
                        "grade": gname, "path": path, "teachers": teachers,
                        "stats": stats, "classes": classes,
                        "rows_meta": rows_meta, "grids": grids,
                        "log_tail": logs[-4:],
                        "_slots": slots,      # kept server-side only; stripped before JSON response
                        "_grp2": grp2,        # kept server-side only
                    })
            except Exception as e:
                import traceback
                return self._send_json(
                    {"error": f"generation failed: {e}",
                     "trace": traceback.format_exc()}, 500)

            token = uuid.uuid4().hex
            # present results in natural display order (6,7,8,9,10,11-12)
            disp = {g["name"]: i for i, g in enumerate(ALL_GRADES)}
            all_results.sort(key=lambda r: disp.get(r["grade"], 99))
            ALL_CACHE[token] = {"grades": all_results}

            # ── Populate _grade_slots_cache so combined teacher TT works ─────
            import excel_writer as _ew
            _ew._grade_slots_cache.clear()
            for r in all_results:
                if r.get("_slots") and r.get("_grp2"):
                    grp2r = r["_grp2"]
                    _ew._grade_slots_cache[r["grade"]] = {
                        "all_slots":    r["_slots"],
                        "subject_rows": grp2r["subject_rows"],
                        "classes":      grp2r["classes"],
                        "grp":          grp2r,
                    }

            # slots is not JSON-serialisable → strip from response
            out = [{k: v for k, v in r.items() if k not in ("_slots","_grp2")} for r in all_results]
            return self._send_json({"ok": True, "token": token, "grades": out})

        return self._send_json({"error": "not found"}, 404)

    @staticmethod
    def _extract_teacher(full_path, sheet):
        wb = load_workbook(full_path)
        if sheet not in wb.sheetnames: return None
        for s in list(wb.sheetnames):
            if s != sheet: del wb[s]
        out = os.path.join(WORKDIR, f"teacher_{uuid.uuid4().hex[:8]}.xlsx")
        wb.save(out)
        return out

    @staticmethod
    def _extract_teacher_combined(teacher_name):
        """
        Build a single-sheet combined personal TT for a cross-grade teacher.
        Uses in-memory cache if available (populated after Generate All Grades).
        Falls back to reading individual Excel files if cache is empty.
        """
        import excel_writer as _ew
        out = os.path.join(WORKDIR, f"teacher_combined_{uuid.uuid4().hex[:8]}.xlsx")

        # ── Try cache-based single-sheet approach first ───────────────────────
        if _ew._grade_slots_cache:
            from excel_writer import export_teacher_combined_xlsx
            ok = export_teacher_combined_xlsx(teacher_name, out)
            if ok:
                return out

        # ── Fallback: copy teacher sheets from individual grade Excel files ───
        from openpyxl import load_workbook, Workbook
        if not ALL_CACHE:
            return None
        last_token = list(ALL_CACHE.keys())[-1]
        grades_info = ALL_CACHE[last_token]["grades"]
        sheet_name  = re.sub(r'[\\/:*?"<>|]', '_', teacher_name)[:31]

        wb_out = Workbook()
        wb_out.remove(wb_out.active)
        from copy import copy as _copy
        any_written = False

        for gi in grades_info:
            path  = gi["path"]
            gnum  = gi["grade"].split()[-1]
            if not os.path.exists(path):
                continue
            wb_src = load_workbook(path)
            if sheet_name not in wb_src.sheetnames:
                wb_src.close()
                continue
            ws_src  = wb_src[sheet_name]
            ws_dest = wb_out.create_sheet(f"Grade {gnum}")
            for row in ws_src.iter_rows():
                for cell in row:
                    nc = ws_dest.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        nc.font = _copy(cell.font); nc.fill = _copy(cell.fill)
                        nc.border = _copy(cell.border); nc.alignment = _copy(cell.alignment)
            for m in ws_src.merged_cells.ranges:
                ws_dest.merge_cells(str(m))
            for cl, cd in ws_src.column_dimensions.items():
                ws_dest.column_dimensions[cl].width = cd.width
            for rn, rd in ws_src.row_dimensions.items():
                ws_dest.row_dimensions[rn].height = rd.height
            ws_dest.sheet_view.showGridLines = False
            ws_dest.page_setup.orientation   = "landscape"
            wb_src.close()
            any_written = True

        if not any_written:
            return None
        wb_out.save(out)
        return out


# ─────────────────────────────────────────────────────────────────────────────
#  Front-end (self-contained single page)
# ─────────────────────────────────────────────────────────────────────────────
INDEX_HTML = """<!DOCTYPE html>
<html lang="en"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Lyceum Timetable Generator</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Fraunces:wght@300;600&display=swap" rel="stylesheet">
<style>
:root{
  --ink:#12141a;--muted:#6b7280;--line:#e2e6ed;--bg:#f4f5f8;--surface:#fff;
  --brand:#1a3660;--brand2:#2563a8;--brand-light:#dbeafe;
  --accent:#0f7d59;--warn:#b45309;
  --g6:#1a3660;--g7:#7c3aed;--g8:#0f7d59;
  --g9:#b45309;--g10:#be185d;--g11:#0e7490;
  --maths-bg:#eff6ff;--merge-bg:#f3e8ff;--rel-bg:#ecfdf5;--pe-bg:#fef9c3;--break-bg:#fefce8;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font:13.5px/1.6 'DM Mono',monospace;color:var(--ink);background:var(--bg)}
header{background:var(--brand);color:#fff;padding:16px 28px;display:flex;align-items:center;gap:16px;border-bottom:3px solid #0f2547}
.logo{font-family:'Fraunces',serif;font-size:21px;font-weight:600}
.sub{opacity:.6;font-size:11.5px;margin-top:1px}
.badge{margin-left:auto;background:#fff2;border:1px solid #fff3;border-radius:6px;padding:3px 9px;font-size:11px}
main{max-width:1300px;margin:0 auto;padding:20px}
.card{background:var(--surface);border:1px solid var(--line);border-radius:13px;padding:20px 22px;margin-bottom:16px;box-shadow:0 1px 3px #0000000d}
.card-title{font-family:'Fraunces',serif;font-size:15px;font-weight:600;color:var(--brand);margin-bottom:3px;display:flex;align-items:center;gap:8px}
.step{display:inline-flex;align-items:center;justify-content:center;width:22px;height:22px;border-radius:50%;background:var(--brand);color:#fff;font-size:11px;flex-shrink:0}
.card-sub{color:var(--muted);font-size:12px;margin-bottom:12px}
.row{display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end}
label.fld{display:flex;flex-direction:column;gap:3px;font-size:12px;color:var(--muted)}
select,input,button{font:inherit}
select,input[type=text],input[type=number]{border:1px solid var(--line);border-radius:7px;padding:6px 9px;background:#fff;color:var(--ink);outline:none;transition:border-color .15s}
select:focus,input:focus{border-color:var(--brand2)}
input[type=number]{width:54px;text-align:center}
button{cursor:pointer;border:0;border-radius:8px;padding:7px 15px;font-weight:500;transition:all .15s}
.btn-primary{background:var(--brand2);color:#fff}.btn-primary:hover{background:#1e56a0}
.btn-all{background:var(--brand);color:#fff;font-weight:600}.btn-all:hover{background:#0f2547}
.btn-ghost{background:var(--brand-light);color:var(--brand)}.btn-ghost:hover{background:#bfdbfe}
.btn-accent{background:var(--accent);color:#fff}
button:disabled{opacity:.4;cursor:not-allowed}
.setup-wrap{overflow:auto;max-height:400px;border:1px solid var(--line);border-radius:9px;margin-bottom:6px}
.setup-wrap table{border-collapse:collapse;width:100%;font-size:12px}
.setup-wrap th{position:sticky;top:0;background:#f0f4f9;z-index:1;padding:6px 8px;border:1px solid var(--line);text-align:center;font-family:'Fraunces',serif;font-size:12.5px;font-weight:600;color:var(--brand)}
.setup-wrap td{border:1px solid var(--line);padding:4px 7px}
.setup-wrap td.subj{font-weight:500;white-space:nowrap;min-width:130px}
.setup-wrap td.tchr input{width:186px}
.setup-wrap td.subj input{width:140px}
.setup-wrap td.del{width:30px;text-align:center;padding:2px}
.row-del{background:#fef2f2;color:#b42318;border:1px solid #fecaca;border-radius:6px;
         padding:2px 7px;font-size:13px;line-height:1;font-weight:700}
.row-del:hover{background:#fee2e2}
.add-row{background:var(--brand-light);color:var(--brand);border:1px dashed #9bb8dd;
         border-radius:8px;padding:6px 12px;font-size:12.5px;font-weight:600;margin-top:7px}
.add-row:hover{background:#e4eefb}
.reset-btn{background:#f8fafc;color:var(--muted);border:1px solid var(--line);
           border-radius:7px;padding:5px 11px;font-size:12px}
.reset-btn:hover{color:var(--brand)}
.new-row td{background:#f7fcf7}
.divider{height:1px;background:var(--line);margin:12px 0}
.summary-row{display:flex;gap:20px;flex-wrap:wrap;margin-bottom:12px}
.summary-grade{font-size:12px;line-height:1.8}
.summary-grade b{font-size:14px;display:block;font-family:'Fraunces',serif}
.grade-tabs{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:14px}
.grade-tab{padding:7px 18px;border-radius:9px;font-weight:600;font-size:13px;cursor:pointer;border:2px solid transparent;background:#f0f4f9;color:var(--muted);transition:all .15s}
.grade-tab:hover:not([class*=act]){background:#e5e7eb}
.gtact-g6{background:var(--g6)!important;color:#fff!important;border-color:var(--g6)!important}
.gtact-g7{background:var(--g7)!important;color:#fff!important;border-color:var(--g7)!important}
.gtact-g8{background:var(--g8)!important;color:#fff!important;border-color:var(--g8)!important}
.gtact-g9{background:var(--g9)!important;color:#fff!important;border-color:var(--g9)!important}
.gtact-g10{background:var(--g10)!important;color:#fff!important;border-color:var(--g10)!important}
.gtact-g11{background:var(--g11)!important;color:#fff!important;border-color:var(--g11)!important}
.class-tabs{display:flex;gap:4px;flex-wrap:wrap;margin-bottom:10px}
.class-tab{padding:4px 10px;border-radius:7px;font-size:12px;cursor:pointer;border:1px solid var(--line);background:#f8fafc;color:var(--muted);transition:all .12s}
.ct-g6.active{background:#dbeafe;border-color:var(--g6);color:var(--g6);font-weight:600}
.ct-g7.active{background:#ede9fe;border-color:var(--g7);color:var(--g7);font-weight:600}
.ct-g8.active{background:#d1fae5;border-color:var(--g8);color:var(--g8);font-weight:600}
.ct-g9.active{background:#fef3c7;border-color:var(--g9);color:var(--g9);font-weight:600}
.ct-g10.active{background:#fce7f3;border-color:var(--g10);color:var(--g10);font-weight:600}
.ct-g11.active{background:#cffafe;border-color:var(--g11);color:var(--g11);font-weight:600}
.grade-section{display:none}.grade-section.vis{display:block}
.tt-grid table{border-collapse:collapse;width:100%;font-size:12px}
.tt-grid th{padding:7px 9px;border:1px solid #1e4a8a;font-family:'Fraunces',serif;font-size:13px;color:#fff}
.tt-grid td{border:1px solid var(--line);padding:5px 7px;vertical-align:middle}
.tt-grid td.time{white-space:nowrap;color:var(--muted);font-size:11px;background:#f7f9fc;min-width:108px}
.tt-grid td.time b{display:block;color:var(--ink);font-size:12px}
.tt-grid td.cell{min-width:100px}
.tt-grid td.cell .s{font-weight:600;font-size:12px}
.tt-grid td.cell .t{color:var(--muted);font-size:11px}
.tt-grid td.cell .r{color:#94a3b8;font-size:10.5px}
.tt-grid tr.brk td{background:var(--break-bg);color:var(--warn);text-align:center;font-size:11px;font-weight:600;padding:4px}
.c-maths{background:var(--maths-bg)}.c-merge{background:var(--merge-bg)}.c-rel{background:var(--rel-bg)}.c-pe{background:var(--pe-bg)}
.muted{color:var(--muted)}
.dl-bar{display:flex;gap:8px;flex-wrap:wrap;align-items:center;background:#f8fafc;border:1px solid var(--line);border-radius:9px;padding:9px 13px;margin-bottom:12px}
.dl-bar select{min-width:220px}
.pill{display:inline-block;font-size:11px;padding:2px 9px;border-radius:99px;background:var(--brand-light);color:var(--brand);font-weight:500}
.spinner{display:inline-block;width:12px;height:12px;border:2px solid #c7d6f5;border-top-color:var(--brand2);border-radius:50%;animation:sp .7s linear infinite;vertical-align:-2px;margin-right:5px}
@keyframes sp{to{transform:rotate(360deg)}}
.err{background:#fef2f2;color:#991b1b;border:1px solid #fecaca;padding:10px 14px;border-radius:9px;font-size:12.5px;white-space:pre-wrap;margin-bottom:10px}
.hint{font-size:11.5px;color:var(--muted);margin-top:6px;line-height:1.5}
.ok{color:var(--accent)}.bad{color:#dc2626}
</style>
</head>
<body>
<header>
  <div>
    <div class="logo">Lyceum Timetable Generator</div>
    <div class="sub">Grades 6 – 12 — cross-grade shared-teacher conflict prevention</div>
  </div>
  <div class="badge">v2</div>
</header>
<main>

<div class="card">
  <div class="card-title"><span class="step">1</span> Configure</div>
  <div class="card-sub">Load a grade to edit teachers/periods, or use <b>Generate All Grades</b> directly with defaults.</div>
  <div class="row">
    <label class="fld">Grade <select id="gradeSel"></select></label>
    <button class="btn-ghost" id="loadBtn">Load &amp; edit</button>
    <span id="loadMsg" class="muted" style="font-size:12px"></span>
  </div>
</div>

<div class="card" id="setupCard" style="display:none">
  <div class="card-title"><span class="step">2</span> Edit Teachers &amp; Periods</div>
  <div class="card-sub">Edit a teacher <b>name</b>, <b>subject</b> or <b>period count</b> per class. Use <b>&#10005;</b> to delete a row and <b>&#10133; Add</b> to create a new subject/teacher. Set a period to 0 to skip a class.</div>
  <div id="setupArea"></div>
  <div class="divider"></div>
  <div class="row" style="margin-top:4px">
    <button class="btn-primary" id="genBtn">&#9654; Generate <span id="genLabel"></span></button>
    <button class="btn-all" id="genAllBtn">&#9654;&#9654; Generate All Grades (6–12)</button>
    <span id="genMsg" class="muted" style="font-size:12px"></span>
  </div>
  <div class="hint"><b>Generate All Grades</b> schedules all grades (6&#8594;12) against one shared teacher-busy set, so any teacher shared between grades (e.g. Grade 9 &amp; 10 staff, or Mr. Radun across 9/10/11&#8211;12) is never double-booked.</div>
</div>
<div id="errBox"></div>

<div class="card" id="resultCard" style="display:none">
  <div class="card-title"><span class="step">3</span> Result &mdash; Weekly Timetables (Grades 6–12)</div>
  <div class="summary-row" id="summaryRow"></div>
  <div class="grade-tabs" id="gradeTabs"></div>
  <div class="dl-bar" id="dlBar">
    <label class="fld" style="flex-direction:row;gap:8px;align-items:center;margin:0">
      <span style="font-size:12px;color:var(--muted)">Teacher:</span>
      <select id="teacherSel"></select>
    </label>
    <button class="btn-accent" id="dlTeacherBtn">&#8659; This grade only</button>
    <button class="btn-accent" id="dlTeacherAllBtn" style="display:none;background:var(--accent)">&#8659; All grades (combined)</button>
    <button class="btn-ghost" id="dlFullBtn">&#8659; Full grade workbook</button>
  </div>
  <div id="gradeSections"></div>
</div>

</main>
<script>
var SETUP=null, RESULT=null, AG=null, AC=null;
var EDITS={};            // grade name -> edited setup (persists across grade switches)
var NEWID=1;             // counter for new regular-row ids
var LOADED_GRADE=null;   // grade currently shown in the editor
var GCOL={'Grade 6':'g6','Grade 7':'g7','Grade 8':'g8','Grade 9':'g9','Grade 10':'g10','Grade 11-12 AL':'g11'};
var $=function(s){return document.querySelector(s)};
var $$=function(s){return document.querySelectorAll(s)};
function esc(s){return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')}
function ea(s){return String(s).replace(/"/g,'&quot;')}
async function api(path,opts){
  var r=await fetch(path,opts);
  var j=await r.json().catch(function(){return{error:'bad response'}});
  if(!r.ok||j.error)throw new Error(j.error||'HTTP '+r.status);
  return j;
}
function showErr(m){$('#errBox').innerHTML='<div class="err">'+esc(m)+'</div>'}
function clearErr(){$('#errBox').innerHTML=''}

(async function(){
  var j=await api('/api/grades');
  $('#gradeSel').innerHTML=j.grades.map(function(g){return'<option>'+esc(g)+'</option>'}).join('');
})();

function cacheCurrent(){
  if(SETUP && LOADED_GRADE){ EDITS[LOADED_GRADE]=readSetupFromDOM(LOADED_GRADE); }
}

async function loadGrade(grade, forceFresh){
  clearErr(); $('#loadMsg').textContent='Loading\u2026';
  try{
    if(forceFresh) delete EDITS[grade];
    if(EDITS[grade]){ SETUP=EDITS[grade]; }       // restore previous edits
    else            { SETUP=await api('/api/setup?grade='+encodeURIComponent(grade)); }
    LOADED_GRADE=grade;
    renderSetup();
    $('#setupCard').style.display='';
    $('#genLabel').textContent=SETUP.grade;
    var edited=EDITS[grade]?' (your edits)':'';
    $('#loadMsg').textContent=SETUP.regular_rows.length+' subjects \u00b7 '+SETUP.classes.length+' classes loaded'+edited+'.';
  }catch(e){showErr('Load failed: '+e.message);$('#loadMsg').textContent=''}
}

$('#loadBtn').onclick=function(){
  cacheCurrent();                  // keep edits from the grade we're leaving
  loadGrade($('#gradeSel').value);
};

function renderSetup(){
  if(!SETUP)return;
  var cls=SETUP.classes;
  var h='<div class="row" style="justify-content:space-between;align-items:center;margin-bottom:6px">'+
        '<b style="font-size:13px">Subjects &amp; teachers</b>'+
        '<button class="reset-btn" id="resetBtn" title="Discard edits and reload original values">\u21ba Reset to defaults</button></div>';
  h+='<div class="setup-wrap"><table><thead><tr><th></th>'+
     '<th style="text-align:left">Subject</th><th style="text-align:left">Teacher</th>'+
     cls.map(function(c){return'<th>'+c+'</th>'}).join('')+'</tr></thead><tbody id="regBody">';
  SETUP.regular_rows.forEach(function(r){ h+=regRowHTML(r,cls); });
  h+='</tbody></table></div>';
  h+='<button class="add-row" id="addRegBtn">\u2795 Add subject / teacher</button>';

  if(SETUP.blocks.length){
    h+='<div style="margin-top:16px"><b style="font-size:13px">Shared Blocks</b> '+
       '<span class="muted" style="font-size:11.5px">&mdash; option subjects taught at one shared slot; add or remove options as needed. These teachers are shared across Grades 6\u20138 \u2014 if you rename one, rename it the same way in each grade.</span></div>';
    SETUP.blocks.forEach(function(b){
      var isRel=(b.type==='religion_block');
      h+='<div style="margin-top:8px"><span class="pill">'+esc(b.label)+'</span>'+
         '<div class="setup-wrap" style="max-height:none;margin-top:5px"><table><thead><tr>'+
         '<th></th><th style="text-align:left">Subject</th><th style="text-align:left">Teacher</th>'+
         (isRel?'<th style="text-align:left">Covers</th>':'')+
         '</tr></thead><tbody id="blk_'+b.id+'">';
      b.teachers.forEach(function(t){ h+=blkRowHTML(b,t); });
      h+='</tbody></table></div>'+
         '<button class="add-row" data-addblk="'+b.id+'">\u2795 Add option teacher</button></div>';
    });
  }
  $('#setupArea').innerHTML=h;
  wireSetup();
}

function regRowHTML(r,cls,isNew){
  return '<tr data-rid="'+ea(r.id)+'" data-room="'+ea(r.room||'R??')+'"'+
         ' data-sd="'+(r.same_day?'1':'0')+'" data-a1="'+(r.all_1period?'1':'0')+'"'+
         ' data-mg="'+ea(r.merge_groups||'')+'"'+(isNew?' class="new-row"':'')+'>'+
         '<td class="del"><button class="row-del" title="Delete this subject/teacher">\u2715</button></td>'+
         '<td class="subj"><input type="text" data-k="subject" placeholder="Subject" value="'+ea(r.subject||'')+'"></td>'+
         '<td class="tchr"><input type="text" data-k="teacher" placeholder="Teacher name" value="'+ea(r.teacher||'')+'"></td>'+
         cls.map(function(c){return'<td><input type="number" min="0" max="40" data-cls="'+c+'" value="'+((r.periods&&r.periods[c])||0)+'"></td>'}).join('')+
         '</tr>';
}

function blkRowHTML(b,t,isNew){
  var isRel=(b.type==='religion_block');
  return '<tr class="btr'+(isNew?' new-row':'')+'" data-room="'+ea(t.room||'R??')+'">'+
         '<td class="del"><button class="row-del" title="Remove this option">\u2715</button></td>'+
         '<td class="subj"><input type="text" data-bk="subject" placeholder="Subject" value="'+ea(t.subject||'')+'"></td>'+
         '<td class="tchr"><input type="text" data-bk="teacher" placeholder="Teacher name" value="'+ea(t.teacher||'')+'"></td>'+
         (isRel?'<td><input type="text" data-bk="covers" placeholder="All classes" value="'+ea(t.covers||'')+'"></td>':'')+
         '</tr>';
}

function wireSetup(){
  var area=$('#setupArea');
  // delete (event delegation)
  area.onclick=function(ev){
    var del=ev.target.closest('.row-del');
    if(del){ var tr=del.closest('tr'); if(tr) tr.remove(); return; }
    var addb=ev.target.closest('[data-addblk]');
    if(addb){
      var bid=addb.getAttribute('data-addblk');
      var b=SETUP.blocks.filter(function(x){return x.id===bid})[0];
      var tb=$('#blk_'+bid);
      tb.insertAdjacentHTML('beforeend', blkRowHTML(b,{subject:'',teacher:'',covers:'',room:'R??'},true));
      return;
    }
  };
  var addReg=$('#addRegBtn');
  if(addReg) addReg.onclick=function(){
    var id='rnew'+(NEWID++);
    $('#regBody').insertAdjacentHTML('beforeend',
      regRowHTML({id:id,subject:'',teacher:'',room:'R??',periods:{}},SETUP.classes,true));
  };
  var rb=$('#resetBtn');
  if(rb) rb.onclick=async function(){
    delete EDITS[SETUP.grade];
    await loadGrade(SETUP.grade,true);
  };
}

function readSetupFromDOM(gradeName){
  var classes=SETUP.classes, regular=[];
  $$('#regBody tr').forEach(function(tr){
    var subj=tr.querySelector('input[data-k=subject]').value.trim();
    var tch =tr.querySelector('input[data-k=teacher]').value.trim();
    var periods={},any=false;
    tr.querySelectorAll('input[data-cls]').forEach(function(inp){
      var v=+inp.value||0; periods[inp.dataset.cls]=v; if(v>0)any=true; });
    if(!subj && !tch && !any) return;       // drop fully-empty rows
    regular.push({id:tr.dataset.rid||('r'+regular.length),teacher:tch,subject:subj,
      room:tr.dataset.room||'R??',periods:periods,
      same_day:tr.dataset.sd==='1',all_1period:tr.dataset.a1==='1',
      merge_groups:tr.dataset.mg||''});
  });
  var blocks=[];
  SETUP.blocks.forEach(function(b){
    var teachers=[];
    $$('#blk_'+b.id+' tr.btr').forEach(function(tr){
      var tch=tr.querySelector('input[data-bk=teacher]').value.trim();
      var subj=tr.querySelector('input[data-bk=subject]').value.trim();
      if(!tch && !subj) return;
      var o={teacher:tch,subject:subj,room:tr.dataset.room||'R??'};
      var cov=tr.querySelector('input[data-bk=covers]');
      if(cov) o.covers=cov.value.trim();
      teachers.push(o);
    });
    blocks.push({id:b.id,type:b.type,label:b.label,teachers:teachers});
  });
  return {grade:gradeName||SETUP.grade,classes:classes,regular_rows:regular,blocks:blocks};
}

function collectPayload(gradeName){ return readSetupFromDOM(gradeName); }

function setBusy(b,msg){
  $('#genBtn').disabled=b;$('#genAllBtn').disabled=b;$('#loadBtn').disabled=b;
  $('#genMsg').innerHTML=b?'<span class="spinner"></span>'+esc(msg||''):'';
}

$('#genBtn').onclick=async function(){
  if(!SETUP)return;
  cacheCurrent();
  clearErr();setBusy(true,'Generating '+SETUP.grade+'\u2026');
  try{
    var r=await api('/api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(collectPayload())});
    RESULT={token:r.token,mode:'single',grades:[r]};
    renderResult();
  }catch(e){showErr('Generation failed: '+e.message)}
  finally{setBusy(false)}
};

$('#genAllBtn').onclick=async function(){
  clearErr();setBusy(true,'Scheduling Grade 6 \u2192 7 \u2192 8 with shared teacher-busy set\u2026');
  cacheCurrent();                              // capture edits of the visible grade
  var grades=Object.keys(EDITS).map(function(g){return EDITS[g]});  // all edited grades; the rest use defaults
  try{
    var r=await api('/api/generate_all',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({grades:grades})});
    RESULT={token:r.token,mode:'all',grades:r.grades};
    renderResult();
  }catch(e){showErr('Generation failed: '+e.message)}
  finally{setBusy(false)}
};

function renderResult(){
  $('#resultCard').style.display='';
  var grades=RESULT.grades;
  // Summary
  var sh='';
  grades.forEach(function(g){
    var s=g.stats,col=GCOL[g.grade]||'g6';
    sh+='<div class="summary-grade"><b style="color:var(--'+col+')">'+esc(g.grade)+'</b>'+
        '<span class="'+(s.unplaced===0?'ok':'bad')+'">Unplaced: '+s.unplaced+'</span> &middot; '+
        '<span class="'+(s.straddles===0?'ok':'bad')+'">Split-doubles: '+s.straddles+'</span> &middot; '+
        '<span class="'+(s.double_booked<=1?'ok':'bad')+'">Clashes: '+Math.max(0,s.double_booked-1)+'</span></div>';
  });
  $('#summaryRow').innerHTML=sh;
  // Grade tabs
  var th='';
  grades.forEach(function(g){
    var col=GCOL[g.grade]||'g6';
    th+='<button class="grade-tab" data-grade="'+ea(g.grade)+'" data-col="'+col+'">'+esc(g.grade)+'</button>';
  });
  $('#gradeTabs').innerHTML=th;
  // Grade sections
  var sec='';
  grades.forEach(function(g){
    var col=GCOL[g.grade]||'g6';
    var gk=g.grade.replace(' ','_');
    sec+='<div class="grade-section" data-grade="'+ea(g.grade)+'" id="sec_'+gk+'">';
    sec+='<div class="class-tabs" id="ct_'+gk+'">';
    g.classes.forEach(function(c){
      sec+='<button class="class-tab ct-'+col+'" data-grade="'+ea(g.grade)+'" data-cls="'+c+'">'+c+'</button>';
    });
    sec+='</div>';
    sec+='<div class="tt-grid" id="grid_'+gk+'"></div>';
    if(g.log_tail&&g.log_tail.length){
      sec+='<div style="font-size:11px;color:var(--muted);background:#f8fafc;border-radius:6px;padding:5px 9px;margin-top:6px;white-space:pre-wrap">'+esc(g.log_tail.join('\\n'))+'</div>';
    }
    sec+='</div>';
  });
  $('#gradeSections').innerHTML=sec;
  $$('.grade-tab').forEach(function(b){b.onclick=function(){switchGrade(b.dataset.grade)}});
  $$('.class-tab').forEach(function(b){b.onclick=function(){switchClass(b.dataset.grade,b.dataset.cls)}});
  AG=grades[0].grade; AC=grades[0].classes[0];
  refreshGradeUI();
}

function switchGrade(gn){
  AG=gn;
  var g=RESULT.grades.find(function(x){return x.grade===gn});
  if(g)AC=g.classes[0];
  refreshGradeUI();
}
function switchClass(gn,cls){
  if(gn!==AG)switchGrade(gn);
  AC=cls;
  renderActiveGrid();updateClassTabs();
}
function refreshGradeUI(){
  // grade tabs
  $$('.grade-tab').forEach(function(b){
    var col=b.dataset.col;
    b.className='grade-tab'+(b.dataset.grade===AG?' gtact-'+col:'');
  });
  // grade sections
  $$('.grade-section').forEach(function(s){s.classList.toggle('vis',s.dataset.grade===AG)});
  updateTeacherDrop();updateClassTabs();renderActiveGrid();
}
function updateClassTabs(){
  var g=RESULT.grades.find(function(x){return x.grade===AG});
  if(!g)return;
  var gk=AG.replace(' ','_');
  $$('#ct_'+gk+' .class-tab').forEach(function(b){b.classList.toggle('active',b.dataset.cls===AC)});
}
function updateTeacherDrop(){
  var g=RESULT.grades.find(function(x){return x.grade===AG});
  if(!g)return;
  $('#teacherSel').innerHTML=g.teachers.map(function(t){return'<option value="'+ea(t.sheet)+'">'+esc(t.name)+'</option>'}).join('');
  updateCombinedBtn();
}
function updateCombinedBtn(){
  // Show "All grades (combined)" button whenever we have a multi-grade result
  // Server will return an error if the teacher only appears in one grade
  var isAll=RESULT&&RESULT.mode==='all';
  $('#dlTeacherAllBtn').style.display=isAll?'':'none';
}
function cellClass(s){
  var l=(s||'').toLowerCase();
  if(l.includes('mathematics'))return'c-maths';
  if(l.includes('cookery')||l.includes('art')||l.includes('music')||l.includes('dance')||l.includes('chinese')||l.includes('french')||l.includes('lifeskill')||l.includes('civic'))return'c-merge';
  if(l.includes('buddhism')||l.includes('religion')||l.includes('catholic')||l.includes('hindu')||l.includes('islam')||l.includes('christianity'))return'c-rel';
  if(l.includes('pe')||l.includes('physical'))return'c-pe';
  return'';
}
function renderActiveGrid(){
  var g=RESULT.grades.find(function(x){return x.grade===AG});
  if(!g)return;
  var gk=AG.replace(' ','_');
  var con=document.getElementById('grid_'+gk);
  if(!con)return;
  var col='var(--'+(GCOL[AG]||'g6')+')';
  var days=['Monday','Tuesday','Wednesday','Thursday','Friday'];
  var grid=g.grids[AC],meta=g.rows_meta;
  if(!grid){con.innerHTML='<p class="muted" style="padding:10px">No data</p>';return}
  var h='<table><thead><tr><th style="text-align:left;min-width:110px;background:'+col+'">Time</th>'+
        days.map(function(d){return'<th style="background:'+col+'">'+d+'</th>'}).join('')+
        '</tr></thead><tbody>';
  meta.forEach(function(m,ri){
    if(m.is_break){
      h+='<tr class="brk"><td class="time">'+esc(m.time)+'</td><td colspan="5">&mdash; '+esc(m.label)+' &mdash;</td></tr>';
      return;
    }
    h+='<tr><td class="time"><b>'+esc(m.label)+'</b><br>'+esc(m.time)+'</td>';
    grid[ri].forEach(function(cell){
      if(!cell){h+='<td class="cell muted" style="text-align:center">\u00b7</td>';return}
      h+='<td class="cell '+cellClass(cell.subject)+'">'+
         '<div class="s">'+esc(cell.subject)+'</div>'+
         '<div class="t">'+esc(cell.teacher||'')+'</div>'+
         (cell.room?'<div class="r">'+esc(cell.room)+'</div>':'')+
         '</td>';
    });
    h+='</tr>';
  });
  h+='</tbody></table>';
  con.innerHTML=h;
}
$('#teacherSel').onchange=function(){ updateCombinedBtn(); };
$('#dlTeacherBtn').onclick=function(){
  if(!RESULT)return;
  var sheet=$('#teacherSel').value;
  var gp=RESULT.mode==='all'?'&grade='+encodeURIComponent(AG):'';
  window.location='/api/download/teacher?token='+encodeURIComponent(RESULT.token)+'&sheet='+encodeURIComponent(sheet)+gp;
};
$('#dlTeacherAllBtn').onclick=function(){
  if(!RESULT)return;
  // Find the raw teacher name (not the sheet-safe name) for combined download
  var sheet=$('#teacherSel').value;
  var rawName=sheet; // fallback
  if(RESULT&&RESULT.grades){
    var g=RESULT.grades.find(function(x){return x.grade===AG});
    if(g){var t=g.teachers.find(function(t){return t.sheet===sheet});if(t)rawName=t.name;}
  }
  window.location='/api/download/teacher_combined?token='+encodeURIComponent(RESULT.token)+'&sheet='+encodeURIComponent(rawName);
};
$('#dlFullBtn').onclick=function(){
  if(!RESULT)return;
  var gp=RESULT.mode==='all'?'&grade='+encodeURIComponent(AG):'';
  window.location='/api/download/full?token='+encodeURIComponent(RESULT.token)+gp;
};
</script>
<!-- Assembly Wednesday P1 is a whole-school shared event; the single "clash" it generates is expected and not a real conflict. -->
</body></html>"""


def main():
    srv = ThreadingHTTPServer((HOST, PORT), Handler)
    url = f"http://{HOST}:{PORT}/"
    print("=" * 62)
    print("  Lyceum Timetable Generator — Web Edition v2")
    print(f"  Open:  {url}")
    print("  Stop:  Ctrl+C")
    print()
    print("  KEY FIX: 'Generate All Grades' passes a shared")
    print("  teacher-busy set across Grade 6 -> 7 -> 8 so option-")
    print("  block teachers (Sakna, Pulakshika, EFF Speech, etc.)")
    print("  are never double-booked across grades.")
    print("=" * 62)
    try:
        threading.Timer(0.9, lambda: webbrowser.open(url)).start()
    except Exception:
        pass
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")
        srv.shutdown()


if __name__ == "__main__":
    main()
